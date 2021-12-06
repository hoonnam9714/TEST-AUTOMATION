import time
import datetime
import sys
import pyautogui
import clipboard
import re
import json
import requests
from bs4 import BeautifulSoup
from account import *
from xpath import *
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait          #웹페이지 로딩 대기
from selenium.webdriver.support import expected_conditions as EC 
from selenium.webdriver.support.select import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import ElementClickInterceptedException
from openpyxl import load_workbook         #엑셀파일 불러오기
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.common.exceptions import JavascriptException

def make_excel():
    sabun = pyautogui.prompt("사용자 사번을 입력해주세요.(ex:2140046)","입력") #사용자 입력에서 가져온 사번 맵핑

    options = webdriver.ChromeOptions()
    options.add_experimental_option("excludeSwitches", ["enable-logging"])
    browser = webdriver.Chrome(options=options)
        
    hmc_url = "https://devhmc.hanwhalife.com:1080/sfa/incomeHmc?etrDvCd=01&token=aa2cb27ab1d84482aced7b5ee21c798f&sfaCmpnDvsn=01&mobUserPrno="+sabun+"&mobUserDvsn=02&offcCode=00000" #HMC QA 테스트 URL
    browser.get(hmc_url) #HMC QA 테스트 URL
    browser.maximize_window()
    browser.implicitly_wait(10)

    custName = pyautogui.prompt("고객명을 입력해주세요.(ex:임의설계, 채경훈)","입력")

    #######################################
    # 고객명 클릭
    #######################################
    elems = browser.find_elements_by_class_name("name") #class_name "name"(고객명)을 가지는 모든 엘리먼트 가져오기
    for elem in elems:    #반복하면서 elems 하나씩 뽑아오기
        if custName in elem.text:
            elem.click()  #일치하는 고객명 클릭
            break # 고객명 클릭 개선 필요

    elems1 = browser.find_elements_by_class_name("af_line") #class_name "af_line"(기설계-상품코드)을 가지는 모든 엘리먼트 가져오기
    time.sleep(0.5)
    goodCode = [] #상품코드 배열
    for e1 in elems1: #기설계 건수 가져오기
        goodCode.append(e1.text.replace("-",""))

    elems1 = browser.find_elements_by_class_name("name_sub") #class_name "name_sub"(기설계-상품명)을 가지는 모든 엘리먼트 가져오기
    time.sleep(0.5)
    goodName = [] #상품명 배열
    count=0
    for e1 in elems1: #기설계 건수 가져오기
        goodName.append(e1.text)
        count = count + 1

    wb = load_workbook("testCase_multi.xlsx") #testCase1.xlsx 파일에서 wb를 불러옴
    ws = wb.active  # 현재 활성화된 sheet 가져옴

    k=1
    i=2 #row
    for k in range(1, count + 1): 
        time.sleep(0.5)
        ws.cell(row=i,column=1).value = "case"+str(i-1) #구분
        ws.cell(row=i,column=2).value = sabun #사번
        ws.cell(row=i,column=3).value = custName #고객명
        ws.cell(row=i,column=4).value = goodCode[k-1] #상품코드
        ws.cell(row=i,column=5).value = goodName[k-1] #상품명

        #######################################
        # 기설계 클릭
        #######################################
        time.sleep(1)
        if count > 1:
            locator = '#app > div > section > div.list_wrap.type02 > ul:nth-child(5) > li.on > div.accordion_content > ul > li:nth-child('+ str(k) +')'
        else:
            locator = '#app > div > section > div.list_wrap.type02 > ul:nth-child(5) > li.on > div.accordion_content > ul > li'
        elem = browser.find_element(By.CSS_SELECTOR, locator)
        browser.execute_script("arguments[0].click();", elem)
        browser.implicitly_wait(10)
        time.sleep(3)

        cttSex = browser.find_element_by_xpath('/html/body/div[2]/div/section/div[2]/dl[2]/dd[1]')
        ws.cell(row=i,column=7).value = cttSex.text[:1] #계약자 성별

        mainSex = browser.find_element_by_xpath('/html/body/div[2]/div/section/div[2]/dl[2]/dd[2]')
        ws.cell(row=i,column=10).value = mainSex.text[:1] #주피 성별

        bogi = browser.find_element_by_xpath('/html/body/div[2]/div/section/div[2]/dl[2]/dd[4]')
        ws.cell(row=i,column=44).value = bogi.text #보험 기간

        nabgi = browser.find_element_by_xpath('/html/body/div[2]/div/section/div[2]/dl[2]/dd[5]')
        ws.cell(row=i,column=45).value = nabgi.text #납입 기간

        # result = browser.execute_script("return document.querySelector('#params').value") #파라미터 전체 가져오기
        # tempInt = result.find("paymCyclCode")+len("paymCyclCode")+3 #납입주기 컬럼 위치 찾기
        # paymCyclCode = result[tempInt:tempInt+2].replace('"','') #납입주기 데이터 자르기 및 "제거
        # print("000 paymCyclCode : " + paymCyclCode)
        
        #######################################
        # 이어서 설계하기 클릭
        #######################################
        browser.find_element_by_class_name("btn_point_bt").click()        #이어서 설계하기
        time.sleep(1)
        browser.implicitly_wait(10)
        time.sleep(2)

        #######################################
        # 보험종류 변경 클릭
        #######################################
        browser.find_element_by_xpath("/html/body/div[2]/div[1]/section/div[2]/div[1]/div/p/button").click()        #보험종류 변경 클릭
        time.sleep(1)

        elems = browser.find_elements_by_class_name("select") #class_name "on"(보험종류)을 가지는 모든 엘리먼트 가져오기
        j=0
        for elem in elems:    #반복하면서 elems 하나씩 뽑아오기
            ws.cell(row=i,column=21+j).value = elem.text #보험종류
            j=j+1
        
        #######################################
        # 적용하기 클릭
        #######################################
        time.sleep(0.5)
        browser.find_element_by_xpath("/html/body/div[2]/div[2]/div[2]/footer/button").click()        #적용하기 클릭
        time.sleep(1)

        ##########################################
        #계약자,주피,종피1/2/3 생년월일 엑셀에 저장
        ##########################################
        elems = browser.find_elements_by_tag_name("input") #input 태그를 모두 가져옴
        birth = []
        l=0
        j=0
        for elem in elems:    #반복하면서 elems 하나씩 뽑아오기
            if elem.get_attribute("title") == "생년월일": #input태그중 title 속성의 값이 "생년월일"
                birth.append(elem.get_attribute("value")) #input값은 개발자도구에서 보이진 않지만 value 속성에 담겨있음!!!
                ws.cell(row=i,column=6+j).value = birth[l] #birth[0] : 계약자생년월일, birth[1] : 주피생년월일, birth[2] : 종피1생년월일...
                l = l+1
                j = j+3

        elems = browser.find_elements_by_class_name("on") #class_name "on"(직종)을 가지는 모든 엘리먼트 가져오기
        j=0
        for elem in elems:    #반복하면서 elems 하나씩 뽑아오기
            if "위험" in elem.text: 
                ws.cell(row=i,column=11+j).value = elem.text #직종
                j = j+3

        try:
            result = browser.execute_script("return document.querySelector('#chk01').checked") #주피 건강체 가져오기
            if result:
                ws.cell(row=i,column=26).value = "Y" 
            else:
                ws.cell(row=i,column=26).value = "N"
        except NoSuchElementException:
            pass
        except JavascriptException:
            pass
            
        ##########################################
        #종피1,2,3 성별/직종 엑셀에 저장
        ##########################################
        try:
            l=2
            j=0
            for l in range(2,5):
                if birth[l]:  #종피1,2,3 생년월일이 존재하면
                    result = browser.execute_script("return document.querySelector('#man0"+str(l)+"').checked") #종피1,2,3 성별 가져오기
                    if result:
                        ws.cell(row=i,column=13+j).value = "남" #종피1,2,3 성별
                    else:
                        ws.cell(row=i,column=13+j).value = "여" #종피1,2,3 성별
                else:
                    ws.cell(row=i,column=13+j).value = "" #종피1,2,3 성별
                    ws.cell(row=i,column=14+j).value = "" #종피1,2,3 직종
                l = l+1
                j = j+3
        except:
            pass

        time.sleep(0.5)
        #######################################
        # 플랜설계 팝업 닫기 
        #######################################
        try:
            elem = browser.find_element_by_xpath('//*[@id="container"]/div[3]/div[2]/div/div/button/span') #직접설계
            browser.execute_script("arguments[0].click();", elem)
        except NoSuchElementException:
            pass

        #######################################
        # 다음버튼 클릭
        #######################################
        browser.implicitly_wait(10)
        time.sleep(0.5)
        browser.find_element_by_xpath("/html/body/div[2]/div[1]/footer/div/button").click()        #적용하기 클릭
        time.sleep(3)

        result = browser.execute_script("return document.querySelector('#plan01').checked") #주계약 가입금액 체크여부
        if result:
            if "연금" in goodName[k-1]:
                ws.cell(row=i,column=28).value = "주계약보험료"
            else:
                ws.cell(row=i,column=28).value = "가입금액" 
        else:
            ws.cell(row=i,column=28).value = "합계보험료"
        
        
        elems = browser.find_elements_by_tag_name("input") #input 태그를 모두 가져옴 (계약금액)
        for elem in elems:    #반복하면서 elems 하나씩 뽑아오기
            if elem.get_attribute("type") == "tel": #input태그중 type 속성의 값이 "tel"
                ws.cell(row=i,column=29).value = elem.get_attribute("value").replace(",","")
                break

        try:
            result = browser.execute_script("return Array.from(document.querySelectorAll('p.s_tit')).find(el => el.textContent === '연금지급형태').nextElementSibling.selectedOptions[0].innerText")
            ws.cell(row=i,column=41).value = result
        except JavascriptException:
            pass

        try:
            result = browser.execute_script("return Array.from(document.querySelectorAll('p.s_tit')).find(el => el.textContent === '연금개시').nextElementSibling.selectedOptions[0].innerText")
            ws.cell(row=i,column=42).value = result
        except JavascriptException:
            pass

        try:
            result = browser.execute_script("return Array.from(document.querySelectorAll('p.s_tit')).find(el => el.textContent === '연금집중기간').nextElementSibling.selectedOptions[0].innerText")
            ws.cell(row=i,column=43).value = result
        except JavascriptException:
            pass

        try:
            result = browser.execute_script("return Array.from(document.querySelectorAll('p.s_tit')).find(el => el.textContent === '납입주기').nextElementSibling.selectedOptions[0].innerText")
            ws.cell(row=i,column=46).value = result.replace(" ","")
        except JavascriptException:
            pass
        
        try:
            result = browser.execute_script("return document.querySelector('#chk02').checked") #환급특약제외 체크여부
            if result:
                ws.cell(row=i,column=47).value = "Y"
            else:
                ws.cell(row=i,column=47).value = "N"
            
            result = browser.execute_script("return document.querySelector('#chk01').checked") #납입면제특약제외 체크여부
            if result:
                ws.cell(row=i,column=48).value = "Y"
            else:
                ws.cell(row=i,column=48).value = "N"
        except NoSuchElementException:
            pass
        except JavascriptException:
            pass

        elems = browser.find_elements_by_tag_name("li") #li 태그를 모두 가져옴(특약코드)
        j=0
        for elem in elems:    #반복하면서 elems(특약코드) 하나씩 뽑아오기
            temp = elem.get_attribute("id")
            if temp and len(temp) == 7:
                ws.cell(row=1,column=49+j).value = "특약코드"
                ws.cell(row=i,column=49+j).value = temp

                ws.cell(row=1,column=50+j).value = "특약명"
                name = browser.find_element_by_xpath('//*[@id='+temp+']/div/p[1]')
                ws.cell(row=i,column=50+j).value = name.text

                ws.cell(row=1,column=51+j).value = "특약금액"
                reslut = browser.find_element_by_xpath('//*[@id='+temp+']/div/p[2]')
                amt = re.sub(r'[^0-9]', '', reslut.text.split('/')[0]) #구분자 '/'로 문자열 자르고 0번째 배열 문자에서 숫자면 가져옴
                amt = amt + "0000"  #50만원 -> 500000원으로 변경... 나중에 검수해서 변경해야 할수도 있음!!!
                ws.cell(row=i,column=51+j).value = amt

                ws.cell(row=1,column=52+j).value = "특약보험기간"

                ws.cell(row=1,column=53+j).value = "특약납입기간"
                period = re.sub(r'[^0-9]', '', reslut.text.split('/')[1]) #구분자 '/'로 문자열 자르고 1번째 배열 문자에서 숫자면 가져옴
                period = period + "년만기"
                ws.cell(row=i,column=53+j).value = period

                j=j+5 #특약관련 컬럼이 5개 존재

        #######################################
        # 상세결과보기 클릭
        #######################################
        time.sleep(0.5)
        reslut = browser.find_element_by_xpath('//*[@id="container"]/div[1]/footer/div/button[2]').click() #상세결과보기
        time.sleep(0.5)
        browser.implicitly_wait(10)
        time.sleep(7)

        #######################################
        # 뒤로가기 클릭
        #######################################
        j=1
        for j in range(1,5):
            browser.find_element_by_xpath('/html/body/div[2]/div[1]/header/button[1]').click() #뒤로가기
            time.sleep(0.5)
            browser.implicitly_wait(10)
            time.sleep(3)
            j=j+1

        #######################################
        # 고객명 클릭
        #######################################
        elems = browser.find_elements_by_class_name("name") #class_name "name"(고객명)을 가지는 모든 엘리먼트 가져오기
        for elem in elems:    #반복하면서 elems 하나씩 뽑아오기
            if custName in elem.text:
                elem.click()  #일치하는 고객명 클릭

        k = k+1
        i = i+1
        # break

    ws.cell(row=1,column=ws.max_column + 1).value = "비고"
    wb.save("testCase_multi.xlsx")
    wb.close()
    browser.quit()

################################################################
# 프로그램 시작점을 구분하기 위해 if __name__ == "__main__": 사용
################################################################
if __name__ == "__main__":
    make_excel()