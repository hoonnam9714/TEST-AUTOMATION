import pyautogui
import pyperclip
import time
import datetime
import requests
from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from datetime import datetime, timedelta      
from bs4 import BeautifulSoup, element  
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys

###############################
# 엑셀 항목 시작 위치
###############################
insrKindNo        = 6       # 보험종류
ctorInfoNo        = 10      # 계약자 정보
basicMainPlanNo   = 35      # 기본설계 주계약
retirePlanNo      = 41      # 연금설계
setAddPayNo       = 43      # 추가설계_추가납입
setDrawNo         = 63      # 추가설계_인출
setSmartChngNo    = 83      # 스마트전환형
setSmartAddPayNo  = 87      # 스마트전환형_추가납입
setPreRetireAmtNo = 107     # 연금선지급
setGuRtChngObjtNo = 110     # 보장전환형
setGuRtAddPayNo   = 116     # 보장전환형_추가납입
setBasicAddPlanNo = 136     # 기본설계_부가특약
setCtorTable_elem = ""
k = 0
l = 0
m = 0
###############################
# 로그인 화면 처리
###############################
def login():
    browser.maximize_window()
    time.sleep(1)
    browser.find_element(By.ID, 'ibxUserId').send_keys(sabun)
    time.sleep(0.5)
    browser.find_element(By.ID, 'ibxPassword').send_keys(password)
    time.sleep(0.5)
    browser.find_element(By.ID, 'rdoTitl_input_0').click() #스마트인증 라디오 버튼
    time.sleep(0.5)
    browser.find_element(By.XPATH, '//*[@id="bt_login"]/a').click() #로그인 버튼
    time.sleep(0.5)
###############################
# 팝업 종료 처리
###############################
def popUpClose():
    browser.implicitly_wait(10)
    time.sleep(1)
    handles = browser.window_handles #메인페이지 및 팝업 갯수

    # 메인페이지가 아니면 팝업(브라우저) 종료
    for handle in handles:
        if handle != handles[0]:
            browser.switch_to.window(handle)
            browser.close()

    time.sleep(0.5)
    browser.switch_to.window(browser.window_handles[0]) #메인화면으로 전환
    elem = browser.find_element(By.XPATH, '//*[@id="smallpop_close"]') #내부팝업(?) 종료
    browser.execute_script("arguments[0].click();", elem)
    time.sleep(0.5)

###############################
# alert 종료 처리
###############################
def alertAccept():
    browser.implicitly_wait(0.5)
    try:
        # 팝업창 처리
        WebDriverWait(browser, 0.5).until(EC.alert_is_present())
        alert = browser.switch_to.alert
        alert.accept() 
    # 팝업창이 없으면
    except:
        pass
###############################
# 엑셀 상품명과 일치하는 메뉴명 찾기
###############################
def goodFind():
    browser.implicitly_wait(0.5)
    browser.switch_to.default_content() #메인화면으로 전환
    browser.find_element(By.ID, 'genMenuDepth1_3_menuNm').click()
    goodName = ws.cell(row=i,column=2).value #엑셀 상품명
    elems = browser.find_elements(By.CLASS_NAME, 'w2anchor2 ') #메뉴명 전체 가져오기
    for elem in elems:
        if str(elem.text) == str(goodName): 
            time.sleep(0.5)
            elem.click()
            print('### ' + str(i) + '. 상품명 : ' + str(goodName))
            break
###############################
# 상품 프레임으로 전환
###############################          
def mainSelect():
    # 메인frame으로 전환
    time.sleep(1)
    browser.switch_to.default_content() #메인화면으로 전환
    time.sleep(0.5)
    browser.switch_to.frame('windowContainer1_subWindow0_iframe') #frame 전환
    #페이지 로딩 대기
    time.sleep(0.5)
    browser.implicitly_wait(10)
    pageWait('//*[@id="sbxMenu0_label"]') #메인 조회 페이지 로딩 대기
###############################
# 인풋값 set
###############################    
def setInput(arg1, arg2):
        elem = browser.find_element(By.ID, arg1)
        time.sleep(0.5)
        elem.clear()
        time.sleep(0.5)
        elem.send_keys(ws.cell(row=i,column=arg2).value)
        time.sleep(0.5)
        pyautogui.press('enter') # 엔터 조회
        time.sleep(0.5)
        browser.implicitly_wait(10)
        pageWait('//*[@id="sbxMenu0_label"]') #메인 조회 페이지 로딩 대기
###############################
# 보험종류 set
###############################
def setInsrKind():
    time.sleep(0.5)
    browser.implicitly_wait(0.5)
    try:
        print("### 보험종류 setInsrKind 시작 ###")
        for k in range(0,4):
            elem = browser.find_element(By.ID, 'sbxMenu'+str(k)+'_label') # 보험종류 selectbox 
            time.sleep(0.5)
            elem.click() ## 보험종류 selectbox click
            if str(elem.text):
                try:
                    for l in range(0,5):
                        elem = browser.find_element(By.ID, 'sbxMenu' + str(k) + '_itemTable_'+ str(l))
                        time.sleep(0.1)
                        if str(elem.text) == str(ws.cell(row=i,column=6+k).value): # 화면값과 엑셀값이 동일하면 클릭
                            elem.click()
                            break
                except Exception as e:
                    print("### 보험종류 selectbox 예외발생: " + str(e))
                    pass
    except Exception as e:
        print("### 보험종류 setInsrKind 예외발생: " + str(e))
        pass
    print("### 보험종류 setInsrKind 종료 ###")
    time.sleep(0.5)
###############################
# 직종정보 set
###############################
def setOccpCode():
    browser.implicitly_wait(0.5)
    global setCtorTable_elem
    global k
    try:
        print("### 직종정보 setOccpCode 시작 ###")
        for k in range(0,5): # 보험계약자, 피보험자(주피), 배우자(종피), 자녀1, 자녀2 row
            setCtorTable_elem = browser.find_element(By.ID, "grdInsu_cell_" + str(k) + "_0") 
            if '보험계약자' == str(setCtorTable_elem.text):
                setOccpCode_detail('보험계약자',ctorInfoNo)
            if '피보험자(주피)' == str(setCtorTable_elem.text):
                setOccpCode_detail('피보험자(주피)',ctorInfoNo+5)
            if '배우자(종피)' == str(setCtorTable_elem.text):
                setOccpCode_detail('배우자(종피)',ctorInfoNo+10)
            if '자녀1' == str(setCtorTable_elem.text):
                setOccpCode_detail('자녀1',ctorInfoNo+15)
            if '자녀2' == str(setCtorTable_elem.text):
                setOccpCode_detail('자녀2',ctorInfoNo+20)
    except Exception as e:
        print("### 직종정보 setOccpCode 예외발생: " + str(e))
        pass
    print("### 직종정보 setOccpCode 종료 ###")
    time.sleep(0.5)   
###############################
# 직종정보_상세 set
###############################
def setOccpCode_detail(arg1,arg2):
    browser.implicitly_wait(0.5)
    global setCtorTable_elem
    global k,l,m
    print("### 직종정보_상세 setOccpCode_detail 시작 ###")
    if arg1 == str(setCtorTable_elem.text):
        # 직종정보
        try:
            # 이름 변경
            selectBoxFind("//*[@id='grdInsu_cell_"+str(k)+"_1']",'G_grdInsu___selectbox_custId_itemTable_',arg2)     
            # 화면 직종정보와 엑셀 직종정보가 다르면
            if str(ws.cell(row=i,column=arg2+2).value) != str(browser.find_element(By.XPATH, "//*[@id='grdInsu_cell_"+str(k)+"_11']/nobr").text):
                print("### 직종정보_상세 setOccpCode_detail 시작 ###")
                try:
                    elem = browser.find_element(By.XPATH, "//*[@id='grdInsu_cell_" + str(k) + "_9']/button")
                    elem.click()
                    pyautogui.press('enter')        # alert창 처리
                    time.sleep(0.5)
                except Exception as e:
                    print("### 직종정보_상세 setOccpCode_detail 직종 팝업창 클릭 및 alert처리 예외발생: " + str(e))
                    pass

                browser.switch_to.frame('uvcmp040pvw_iframe') #frame 전환
                try:
                    #페이지 로딩 대기
                    time.sleep(0.5)
                    browser.implicitly_wait(10)
                    pageWait('//*[@id="ibxSrchWord"]') #직종 조회 페이지 로딩 대기
                    elem = browser.find_element(By.ID, 'ibxSrchWord')
                    elem.send_keys(ws.cell(row=i,column=arg2+2).value)
                    elem.send_keys(Keys.ENTER)
                    time.sleep(2)
                except Exception as e:
                    print("### 직종정보_상세 setOccpCode_detail 직종코드 클릭 예외발생: " + str(e))
                    pass
                try:
                    elem = browser.find_element(By.XPATH, '//*[@id="grdBsns_cell_0_0"]/nobr')
                    elem.click()
                    time.sleep(0.5)
                except Exception as e:
                    print("### 직종정보_상세 setOccpCode_detail 직종명 클릭 예외발생: " + str(e))
                    pass
                try:
                    pageWait('//*[@id="btnConfirm"]/a') #선택 로딩 대기
                    elem = browser.find_element(By.XPATH, '//*[@id="btnConfirm"]/a')
                    elem.click()
                    time.sleep(5)
                    pyautogui.press('enter')        # alert창 처리
                    time.sleep(0.5)
                except Exception as e:
                    print("### 직종정보_상세 setOccpCode_detail 선택 클릭 예외발생: " + str(e))
                    pass
                mainSelect() # 메인화면 프레임 전환 및 대기
        except Exception as e:
            print("### 직종정보_상세 setOccpCode_detail 직종 선택 예외발생: " + str(e))
            pass
    print("### 직종정보_상세 setOccpCode_detail 종료 ###")
    time.sleep(0.5)
###############################
# 계약자정보 set
###############################
def setCtorTable():
    browser.implicitly_wait(0.5)
    global setCtorTable_elem
    global k
    try:
        print("### 계약자정보 setCtorTable 시작 ###")
        for k in range(0,5): # 보험계약자, 피보험자(주피), 배우자(종피), 자녀1, 자녀2 row
            setCtorTable_elem = browser.find_element(By.ID, "grdInsu_cell_" + str(k) + "_0") 
            if '보험계약자' == str(setCtorTable_elem.text):
                setCtorTable_detail('보험계약자',ctorInfoNo)
            if '피보험자(주피)' == str(setCtorTable_elem.text):
                setCtorTable_detail('피보험자(주피)',ctorInfoNo+5)
            if '배우자(종피)' == str(setCtorTable_elem.text):
                setCtorTable_detail('배우자(종피)',ctorInfoNo+10)
            if '자녀1' == str(setCtorTable_elem.text):
                setCtorTable_detail('자녀1',ctorInfoNo+15)
            if '자녀2' == str(setCtorTable_elem.text):
                setCtorTable_detail('자녀2',ctorInfoNo+20)
    except Exception as e:
        print("### 계약자정보 setCtorTable 예외발생: " + str(e))
        pass
    print("### 계약자정보 setCtorTable 종료 ###")
    time.sleep(0.5)
###############################
# 계약자정보_상세 set
###############################
def setCtorTable_detail(arg1,arg2):
    browser.implicitly_wait(0.5)
    global setCtorTable_elem
    global k,l,m
    print("### 계약자정보_상세 setCtorTable_detail 시작 ###")
    if arg1 == str(setCtorTable_elem.text):
        # 이름
        selectBoxFind("//*[@id='grdInsu_cell_"+str(k)+"_1']",'G_grdInsu___selectbox_custId_itemTable_',arg2)           
        # 건강체
        try:
            elem = browser.find_element(By.ID, 'grdInsu_cell_' + str(k) + '_8') # 건강체 check box
            checked = browser.find_element(By.XPATH, "//*[@id='grdInsu_cell_"+str(k)+"_8']/input")
            if 'Y' == str(ws.cell(row=i,column=arg2+1).value):
                try:
                    if checked.get_attribute('checked'):
                        pass
                    else:
                        elem.click() # 건강체 선택
                except:
                    elem.click() # 건강체 선택
            else:
                try:
                    if checked.get_attribute('checked'):
                        elem.click() # 건강체 선택 취소
                    else:
                        pass
                except:
                    pass
        except Exception as e:
            print("### 계약자정보_상세 setCtorTable_detail 건강체 선택 예외발생: " + str(e))
            pass
        # 외국인  
        selectBoxFind("//*[@id='grdInsu_cell_"+str(k)+"_15']",'G_grdInsu___selectbox_custFrnrDvsn_itemTable_',arg2+3)
        # 체류 
        selectBoxFind("//*[@id='grdInsu_cell_"+str(k)+"_16']",'G_grdInsu___selectbox_custFrnr_itemTable_',arg2+4)
    print("### 계약자정보_상세 setCtorTable_detail 종료 ###")
    time.sleep(0.5)
###############################
# 기본설계 주계약 set
###############################
def setBasicMainPlan():
    print("### 기본설계 주계약 setBasicMainPlan 시작 ###")
    browser.implicitly_wait(0.5)
    alertAccept() # alert 처리  
    try:
        elem = browser.find_element(By.XPATH ,'//*[@id="grdMain_cell_0_1"]/nobr') #합계보험료 표
        time.sleep(0.5)
        elem.click()
        time.sleep(0.5)
        elem = browser.find_element(By.ID, 'G_grdMain__totPrem') #합계보험료 input
        elem.send_keys(ws.cell(row=i,column=basicMainPlanNo).value)
        time.sleep(0.5)
    except Exception as e:
        print("### 기본설계 주계약 setBasicMainPlan 합계보험료 입력 예외발생: " + str(e))
        pass
    try:
        elem = browser.find_element(By.XPATH, '//*[@id="grdMain_cell_0_4"]/nobr') #가입금액 표
        time.sleep(0.5)
        elem.click()
        time.sleep(0.5)
        elem = browser.find_element(By.ID, 'G_grdMain__amt') #가입금액 input
        elem.send_keys(ws.cell(row=i,column=basicMainPlanNo+1).value) 
    except Exception as e:
        print("### 기본설계 주계약 setBasicMainPlan 가입금액 입력 예외발생: " + str(e))
        pass
    #보험기간
    selectBoxFind('//*[@id="grdMain_cell_0_7"]/nobr','G_grdMain___selectbox_intr_itemTable_itemTable_',basicMainPlanNo+2)
    #납입기간
    selectBoxFind('//*[@id="grdMain_cell_0_9"]/nobr','G_grdMain___selectbox_pytr_itemTable_',basicMainPlanNo+3)
    #납입주기
    selectBoxFind('//*[@id="grdMain_cell_0_11"]/nobr','G_grdMain___selectbox_paym_itemTable_',basicMainPlanNo+4)
    #증액보험기간
    selectBoxFind('//*[@id="grdMain_cell_0_13"]/nobr','G_grdMain___selectbox_icco_itemTable_',basicMainPlanNo+5)

    print("### 기본설계 주계약 setBasicMainPlan 종료 ###")
    time.sleep(0.5)
###############################
# 연금설계 set
###############################
def setRetirePlan():
    print("### 연금설계 setRetirePlan 시작 ###")
    browser.implicitly_wait(0.5)
    alertAccept() # alert 처리
    try:
        # 연금설계탭
        elem = browser.find_element(By.XPATH, '//*[@id="subTac_tab_tab02"]/div[1]/a') 
        elem.click()
        if str(ws.cell(row=1,column=retirePlanNo).value) == '연금설계':
            if ws.cell(row=i,column=retirePlanNo).value:
                #연금개시나이
                selectBoxFind('//*[@id="sbxAnntStrtAge_label"]','sbxAnntStrtAge_itemTable_',retirePlanNo)
                #예시나이간격
                selectBoxFind('//*[@id="sbxAnntAgeTerm_label"]','sbxAnntAgeTerm_itemTable_',retirePlanNo+1)
    except Exception as e:
        print("### 연금설계 setRetirePlan 연금개시나이 예외발생: " + str(e))
        pass
    time.sleep(0.5)
    print("### 연금설계 setRetirePlan 종료 ###")
###############################
# 추가설계_추가납입 set
###############################
def setAddPay():
    print("### 추가설계_추가납입 setAddPay 시작 ###")
    browser.implicitly_wait(0.5)
    alertAccept() # alert 처리
    # 추가설계
    try:
        # 추가설계탭
        elem = browser.find_element(By.XPATH, '//*[@id="subTac_tab_tab03"]/div[1]/a') 
        elem.click()
        temp = 0
        for k in range(0,5):
            if str(ws.cell(row=1,column=setAddPayNo+temp).value) == '추가설계_추가납입':
                if ws.cell(row=i,column=setAddPayNo+temp).value:
                    elem = browser.find_element(By.XPATH, '//*[@id="btnGurtAddPayAddRow"]/a') 
                    elem.click()
                    time.sleep(0.1)
                    # 추가납입시점
                    selectBoxFind("//*[@id='grdGurtAddPay_cell_"+str(k)+"_0']/nobr",'G_grdGurtAddPay___selectbox_paymOrdrFrom_itemTable_',setAddPayNo+temp)
                    # 추가납입기간
                    selectBoxFind("//*[@id='grdGurtAddPay_cell_"+str(k)+"_2']/nobr",'G_grdGurtAddPay___selectbox_paymOrdrTo_itemTable_',setAddPayNo+temp+1)
                    # 추가납입주기
                    selectBoxFind("//*[@id='grdGurtAddPay_cell_"+str(k)+"_4']/nobr",'G_grdGurtAddPay___selectbox_paymUnit_itemTable_',setAddPayNo+temp+2)
                    # 추가납입금액
                    try:
                        elem = browser.find_element(By.XPATH, "//*[@id='grdGurtAddPay_cell_"+str(k)+"_5']/nobr")
                        elem.click()
                        elem = browser.find_element(By.XPATH, '//*[@id="G_grdGurtAddPay__addPaymAmt"]')
                        elem.send_keys(ws.cell(row=i,column=setAddPayNo+temp+3).value)
                    except Exception as e:
                        print("### 추가설계 setAddPay 추가납입금액 예외발생: " + str(e))
                        pass
            temp = temp + 4
    except Exception as e:
        print("### 추가설계 setAddPay 예외발생: " + str(e))
        pass
        print("### 추가설계_추가납입 setAddPay 종료 ###")
        time.sleep(0.5)
###############################
# 추가설계_인출 set
###############################
def setDraw():
    print("### 추가설계_인출 setDraw 시작 ###")
    browser.implicitly_wait(0.5)
    alertAccept() # alert 처리
    # 추가설계
    try:
        # 추가설계탭
        elem = browser.find_element(By.XPATH, '//*[@id="subTac_tab_tab03"]/div[1]/a') 
        elem.click()
        temp = 0
        for k in range(0,5):
            if str(ws.cell(row=1,column=setDrawNo+temp).value) == '추가설계_인출':
                if ws.cell(row=i,column=setDrawNo+temp).value:
                    elem = browser.find_element(By.XPATH, '//*[@id="btnGurtDrawAddRow"]/a') 
                    elem.click()
                    time.sleep(0.5)
                    # 용도  
                    try: 
                        elem = browser.find_element(By.XPATH, "//*[@id='grdGurtDraw_cell_"+str(k)+"_0']/nobr")
                        elem.click()
                        elem = browser.find_element(By.XPATH, '//*[@id="G_grdGurtDraw__drawUseTitle"]')
                        elem.send_keys(str(ws.cell(row=i,column=setDrawNo+temp).value))
                    except Exception as e:
                        print("### 추가설계_인출 setDraw 용도 예외발생: " + str(e))
                        pass
                    # 인출시기_시작
                    selectBoxFind("//*[@id='grdGurtDraw_cell_"+str(k)+"_1']/nobr",'G_grdGurtDraw___selectbox_drawStartYear_itemTable_',setDrawNo+temp+1)
                    # 인출시기_종료
                    selectBoxFind("//*[@id='grdGurtDraw_cell_"+str(k)+"_4']/nobr",'G_grdGurtDraw___selectbox_drawTerm_itemTable_',setDrawNo+temp+2)
                    # 인출금액
                    try:
                        elem = browser.find_element(By.XPATH, "//*[@id='grdGurtDraw_cell_"+str(k)+"_6']/nobr")
                        elem.click()
                        elem = browser.find_element(By.XPATH, '//*[@id="G_grdGurtDraw__drawAmt"]')
                        elem.send_keys(ws.cell(row=i,column=setDrawNo+temp+3).value)
                    except Exception as e:
                        print("### 추가설계_인출 setDraw 인출금액 예외발생: " + str(e))
                        pass
            temp = temp + 4
    except Exception as e:
        print("### 추가설계 setDraw 예외발생: " + str(e))
        pass
    print("### 추가설계_인출 setDraw 종료 ###")
    time.sleep(0.5)
###############################
# 스마트전환형 set
###############################
def setSmartChng():
    print("### 스마트전환형 setSmartChng 시작 ###")
    browser.implicitly_wait(0.5)
    alertAccept() # alert 처리
    try:
        # 스마트전환형탭
        elem = browser.find_element(By.XPATH, '//*[@id="subTac_tab_tab05"]/div[1]/a') #스마트전환형탭
        if '스마트전환형' == str(elem.text): 
            elem.click()
        if str(ws.cell(row=1,column=setSmartChngNo).value) == '스마트전환형':
            if ws.cell(row=i,column=setSmartChngNo).value:
                # 스마트전환 대상자
                selectBoxFind("//*[@id='sbxConv8GurtConv_label']",'sbxConv8GurtConv_itemTable_',setSmartChngNo)
                # 스마트전환 시점
                selectBoxFind("//*[@id='sbxConv1GurtConv_label']",'sbxConv1GurtConv_itemTable_',setSmartChngNo+1)
                # 스마트전환 보험기간
                selectBoxFind("//*[@id='sbxConv4GurtConv_label']",'sbxConv4GurtConv_itemTable_',setSmartChngNo+2)
                # 스마트전환 납입기간
                selectBoxFind("//*[@id='sbxConv5GurtConv_label']",'sbxConv5GurtConv_itemTable_',setSmartChngNo+3)
    except Exception as e:
        print("### 스마트전환형 setSmartChng 예외발생: " + str(e))
        pass
    print("### 스마트전환형 setSmartChng 종료 ###")
    time.sleep(0.5)
###############################
# 스마트전환형_추가납입 set
###############################
def setSmartAddPay():
    print("### 스마트전환형_추가납입 setSmartAddPay 시작 ###")
    browser.implicitly_wait(0.5)
    alertAccept() # alert 처리
    # 스마트전환형_추가납입
    try:
        # 스마트전환형탭
        elem = browser.find_element(By.XPATH, '//*[@id="subTac_tab_tab05"]/div[1]/a') #스마트전환형탭
        elem.click()
        temp = 0
        for k in range(0,5):
            if str(ws.cell(row=1,column=setSmartAddPayNo+temp).value) == '스마트전환형_추가납입':
                if ws.cell(row=i,column=setSmartAddPayNo+temp).value:
                    elem = browser.find_element(By.XPATH, '//*[@id="btnGurtConvAddPayAddRow"]/a') 
                    elem.click()
                    time.sleep(0.1)
                    # 추가납입시점
                    selectBoxFind("//*[@id='grdGurtConvAddPay_cell_"+str(k)+"_0']/nobr",'G_grdGurtConvAddPay___selectbox_paymOrdrFrom_itemTable_',setSmartAddPayNo+temp)
                    # 추가납입기간
                    selectBoxFind("//*[@id='grdGurtConvAddPay_cell_"+str(k)+"_2']/nobr",'G_grdGurtConvAddPay___selectbox_paymOrdrTo_itemTable_',setSmartAddPayNo+temp+1)
                    # 추가납입주기
                    selectBoxFind("//*[@id='grdGurtConvAddPay_cell_"+str(k)+"_4']/nobr",'G_grdGurtConvAddPay___selectbox_paymUnit_itemTable_',setSmartAddPayNo+temp+2)
                    # 추가납입금액
                    try:
                        elem = browser.find_element(By.XPATH, "//*[@id='grdGurtConvAddPay_cell_"+str(k)+"_5']/nobr")
                        elem.click()
                        elem = browser.find_element(By.XPATH, '//*[@id="G_grdGurtConvAddPay__addPaymAmt"]')
                        elem.send_keys(ws.cell(row=i,column=setSmartAddPayNo+temp+3).value)
                    except Exception as e:
                        print("### 스마트전환형_추가납입 setSmartAddPay 추가납입금액 예외발생: " + str(e))
                        pass
            temp = temp + 4
    except Exception as e:
        print("### 스마트전환형_추가납입 setSmartAddPay 예외발생: " + str(e))
        pass
    print("### 스마트전환형_추가납입 setSmartAddPay 종료 ###")
    time.sleep(0.5)
###############################
# 연금선지급 set
###############################
def setPreRetireAmt():
    print("### 연금선지급 setPreRetireAmt 시작 ###")
    browser.implicitly_wait(0.5)
    alertAccept() # alert 처리
    try:
        # 연금선지급탭
        elem = browser.find_element(By.XPATH, '//*[@id="subTac_tab_tab07"]/div[1]/a') 
        elem.click()
        # 데이터 스크래핑
        if str(ws.cell(row=1,column=setPreRetireAmtNo).value) == '연금선지급':
            if ws.cell(row=i,column=setPreRetireAmtNo).value:
                #선지급나이
                selectBoxFind('//*[@id="sbxAnntAdvpymt1_label"]','sbxAnntAdvpymt1_itemTable_',setPreRetireAmtNo)
                #선지급비율
                selectBoxFind('//*[@id="sbxAnntAdvpymt2_label"]','sbxAnntAdvpymt2_itemTable_',setPreRetireAmtNo+1)
                #선지급기간
                selectBoxFind('//*[@id="sbxAnntAdvpymt3_label"]','sbxAnntAdvpymt3_itemTable_',setPreRetireAmtNo+2)
    except Exception as e:
        print("### 연금선지급 setPreRetireAmt 예외발생: " + str(e))
        pass
    print("### 연금선지급 setPreRetireAmt 종료 ###")
    time.sleep(0.5)
###############################
# 보장전환형_대상 set
###############################
def setGuRtChngObjt():
    print("### 보장전환형_대상 setGuRtChngObjt 시작 ###")
    browser.implicitly_wait(0.5)
    alertAccept() # alert 처리
    try:
        # 보장전환형탭
        elem = browser.find_element(By.XPATH, '//*[@id="subTac_tab_tab05"]/div[1]/a')
        if '보장전환형' == elem.text: 
            elem.click()
        if str(ws.cell(row=1,column=setGuRtChngObjtNo).value) == '보장전환형_대상':
            if ws.cell(row=i,column=setGuRtChngObjtNo).value:
                # 보장전환형_대상
                selectBoxFind('//*[@id="sbxConv8GurtConv_label"]','sbxConv8GurtConv_itemTable_',setGuRtChngObjtNo)
                # 보장전환형 시점
                selectBoxFind('//*[@id="sbxConv1GurtConv_label"]','sbxConv1GurtConv_itemTable_',setGuRtChngObjtNo+1)
                # 보장전환형 보험기간
                selectBoxFind('//*[@id="sbxConv4GurtConv_label"]','sbxConv4GurtConv_itemTable_',setGuRtChngObjtNo+2)
                # 보장전환형 납입기간
                selectBoxFind('//*[@id="sbxConv5GurtConv_label"]','sbxConv5GurtConv_itemTable_',setGuRtChngObjtNo+3)
    except Exception as e:
        print("### 보장전환형_대상 setGuRtChngObjt 예외발생: " + str(e))
        pass
    print("### 보장전환형_대상 setGuRtChngObjt 종료 ###")
    time.sleep(0.5)
###############################
# 보장전환형_연금전환설계 set
###############################
def setGuRtRetireChng():
    print("### 보장전환형_연금전환설계 setGuRtRetireChng 시작 ###")
    browser.implicitly_wait(0.5)
    alertAccept() # alert 처리
    try:
        # 보장전환형탭
        elem = browser.find_element(By.XPATH, '//*[@id="subTac_tab_tab05"]/div[1]/a')
        if '보장전환형' == elem.text: 
            elem.click()
        if str(ws.cell(row=1,column=setGuRtChngObjtNo+4).value) == '보장전환형_연금전환설계':
            if ws.cell(row=i,column=setGuRtChngObjtNo+4).value:
                # 보장전환형_연금전환설계_연금개시나이
                selectBoxFind('//*[@id="sbxGurtConvAnntStrtAge_label"]','sbxGurtConvAnntStrtAge_itemTable_',setGuRtChngObjtNo+4)
                # 보장전환형_연금전환설계_예시나이간격
                selectBoxFind('//*[@id="sbxGurtConvAnntAgeTerm_label"]','sbxGurtConvAnntAgeTerm_itemTable_',setGuRtChngObjtNo+5)
    except Exception as e:
        print("### 보장전환형_연금전환설계 setGuRtRetireChng 예외발생: " + str(e))
        pass
    print("### 보장전환형_연금전환설계 setGuRtRetireChng 종료 ###")
    time.sleep(0.5)
###############################
# 보장전환형_추가납입 set
###############################
def setGuRtAddPay():
    print("### 보장전환형_추가납입 getGuRtAddPay 시작 ###")
    browser.implicitly_wait(0.5)
    alertAccept() # alert 처리
    try:
        # 보장전환형탭
        elem = browser.find_element(By.XPATH, '//*[@id="subTac_tab_tab05"]/div[1]/a')
        if '보장전환형' == elem.text: 
            elem.click()
        temp = 0
        for k in range(0,5):
            if str(ws.cell(row=1,column=setGuRtAddPayNo+temp).value) == '보장전환형_추가납입':
                if ws.cell(row=i,column=setGuRtAddPayNo+temp).value:
                    elem = browser.find_element(By.XPATH, '//*[@id="btnGurtConvAddPayAddRow"]/a') 
                    elem.click()
                    time.sleep(0.1)
                    # 추가납입시점    
                    selectBoxFind("//*[@id='grdGurtConvAddPay_cell_"+str(k)+"_0']/nobr",'G_grdGurtConvAddPay___selectbox_paymOrdrFrom_itemTable_',setGuRtAddPayNo+temp)
                    # 추가납입기간
                    selectBoxFind("//*[@id='grdGurtConvAddPay_cell_"+str(k)+"_2']/nobr",'G_grdGurtConvAddPay___selectbox_paymOrdrTo_itemTable_',setGuRtAddPayNo+temp+1)
                    # 추가납입주기
                    selectBoxFind("//*[@id='grdGurtConvAddPay_cell_"+str(k)+"_4']/nobr",'G_grdGurtConvAddPay___selectbox_paymUnit_itemTable_',setGuRtAddPayNo+temp+2)
                    # 추가납입금액
                    try:
                        elem = browser.find_element(By.XPATH, "//*[@id='grdGurtConvAddPay_cell_"+str(k)+"_5']/nobr")
                        elem.click()
                        elem = browser.find_element(By.XPATH, '//*[@id="G_grdGurtConvAddPay__addPaymAmt"]')
                        elem.send_keys(ws.cell(row=i,column=setGuRtAddPayNo+temp+3).value)
                    except Exception as e:
                        print("### 보장전환형_추가납입 setGuRtAddPay 추가납입금액 예외발생: " + str(e))
                        pass
            temp = temp + 4
    except Exception as e:
        print("### 보장전환형_추가납입 setGuRtAddPay 예외발생: " + str(e))
        pass
    print("### 보장전환형_추가납입 setGuRtAddPay 종료 ###")
    time.sleep(0.5)
###############################
# 기본설계_부가특약 가져오기
###############################
def setBasicAddPlan():
    browser.implicitly_wait(0.5)
    try:
        elem = browser.find_element(By.XPATH, '//*[@id="subTac_tab_tab01"]/div[1]/a') #기본설계탭
        elem.click()
    
        # 부가특약 1~9라인 처리
        temp =0
        for row in range(0,9):
            browser.implicitly_wait(0.5)
            elem = browser.find_element(By.XPATH ,"//*[@id='grdGoodMenu_cell_" + str(row) + "_1']/nobr") 
            #특약명칭
            if str(elem.text) == str(ws.cell(row=i,column=setBasicAddPlanNo+temp+1).value):
                #보험기간
                selectBoxFind("//*[@id='grdGoodMenu_cell_" + str(row) + "_2']/nobr",'G_grdGoodMenu___selectbox_goodIntr_itemTable_',setBasicAddPlanNo+temp+2)
                #납입기간
                selectBoxFind("//*[@id='grdGoodMenu_cell_" + str(row) + "_3']/nobr",'G_grdGoodMenu___selectbox_goodPytr_itemTable_',setBasicAddPlanNo+temp+3)
                #가입금액  //*[@id="G_grdGoodMenu__goodPrem"]
                try:
                    elem = browser.find_element(By.XPATH, "//*[@id='grdGoodMenu_cell_" + str(row) + "_4']/nobr") 
                    elem.click()
                    elem = browser.find_element(By.XPATH, "//*[@id='G_grdGoodMenu__goodPrem']")
                    elem.send_keys(str(ws.cell(row=i,column=setBasicAddPlanNo+temp+4).value))
                except Exception as e:
                    print("### 기본설계_부가특약 setBasicAddPlan 1~9라인 가입금액 예외발생: " + str(e))
                    pass
                temp = temp + 5
        
        #10번째 라인부터 동적스크롤 처리
        while True:
            # 현재 9번째 라인 특약명칭
            current_elem = browser.find_element(By.XPATH , '//*[@id="grdGoodMenu_cell_8_1"]/nobr') 
            current_elem.click()
            current9line = str(current_elem.text)

            # 작은 스크롤 아래로 한칸 이동
            current_elem.click()
            time.sleep(0.1)
            pyautogui.press('down') 
            time.sleep(0.1)

            # 동적스크롤 작동 후 9번째 라인 특약명칭
            next_elem = browser.find_element(By.XPATH , '//*[@id="grdGoodMenu_cell_8_1"]/nobr') 
            next9line = str(next_elem.text)

            if str(current9line) == str(next9line): #현재 특약명과 다음 특약명이 일치하면 종료
                break
            else:
                if str(next9line) == str(ws.cell(row=i,column=setBasicAddPlanNo+temp+1).value):
                    time.sleep(0.1)
                    #보험기간 
                    try:
                        selectBoxFind("//*[@id='grdGoodMenu_cell_8_2']/nobr",'G_grdGoodMenu___selectbox_goodIntr_itemTable_',setBasicAddPlanNo+temp+2)
                    except Exception as e:
                        print("### 기본설계_부가특약 setBasicAddPlan 동적스크롤 보험기간 예외발생: " + str(e))
                        pass
                    time.sleep(0.1)
                    #납입기간
                    try:
                        selectBoxFind("//*[@id='grdGoodMenu_cell_8_3']/nobr",'G_grdGoodMenu___selectbox_goodPytr_itemTable_',setBasicAddPlanNo+temp+3)
                    except Exception as e:
                        print("### 기본설계_부가특약 setBasicAddPlan 동적스크롤 납입기간 예외발생: " + str(e))
                        pass
                    time.sleep(0.1)
                    #가입금액
                    try:
                        elem = browser.find_element(By.XPATH, "//*[@id='grdGoodMenu_cell_8_4']/nobr")
                        elem.click()
                        elem = browser.find_element(By.XPATH, "//*[@id='G_grdGoodMenu__goodPrem']")
                        elem.send_keys(str(ws.cell(row=i,column=setBasicAddPlanNo+temp+4).value))
                    except Exception as e:
                        print("### 기본설계_부가특약 setBasicAddPlan 동적스크롤 가입금액 예외발생: " + str(e))
                        pass
                    temp = temp + 5
    except Exception as e:
        print("### 기본설계_부가특약 setBasicAddPlan 예외발생: " + str(e))
        pass
    print("### 기본설계_부가특약 setBasicAddPlan 종료 ###")
    time.sleep(0.5)
###############################
# 가입설계 저장
###############################
def joinSave():
    elem = browser.find_element(By.XPATH,'//*[@id="btnImgPrev"]/a') #결과보기
    elem.click() 
    time.sleep(0.5)
    #가입설계 TIP frame
    try:
        time.sleep(2)
        browser.switch_to.frame('ncjsi456pvw_iframe') #frame 전환
        browser.implicitly_wait(1)
        pageWait('//*[@id="btnJoinPlan"]/a')
        elem = browser.find_element(By.XPATH, '//*[@id="btnJoinPlan"]/a')
        elem.click()
    except Exception as e:
        print("### 가입설계 TIP frame 예외발생: " + str(e))
        pass
    # 결과보기 팝업화면으로 창 전환
    try:
        handles = browser.window_handles
        time.sleep(3)
        browser.switch_to.window(browser.window_handles[1]) 
        time.sleep(3)
    except Exception as e:
        print("### 결과보기 팝업화면으로 창 전환 예외발생: " + str(e))
        pass
    #저장
    try:
        browser.implicitly_wait(10)
        pageWait('//*[@id="btnBottom6"]/a')
        elem = browser.find_element(By.XPATH, '//*[@id="btnBottom6"]/a') 
        browser.execute_script("arguments[0].click();", elem)
        time.sleep(3)
    except Exception as e:
        print("### 저장 예외발생: " + str(e))
        pass
    #닫기
    try:
        browser.implicitly_wait(10)
        alertAccept() # alert 처리
        pageWait('//*[@id="btnClose"]/a')
        elem = browser.find_element(By.XPATH, '//*[@id="btnClose"]/a') 
        browser.execute_script("arguments[0].click();", elem)
    except Exception as e:
        print("### 닫기 예외발생: " + str(e))
        pass
    browser.switch_to.window(browser.window_handles[0]) #메인화면으로 전환
##########################################################################################
# selectBox 엘리먼트 매핑 arg1: selectbox XPATH, arg2: selectBox option ID, arg3: 엑셀시작위치
##########################################################################################
def selectBoxFind(arg1,arg2,arg3):
    try:
        elem = browser.find_element(By.XPATH,arg1)
        elem.click()
        for k in range(0,100):
            elem = browser.find_element(By.ID, arg2+str(k))
            if str(elem.text) == str(ws.cell(row=i,column=arg3).value):
                elem.click()
                break
    except Exception as e:
        print("### selectBoxFind() 예외발생: " + str(e))
        pass
###############################
# 페이지 로딩 대기
###############################
def pageWait(arg1):
    elem = WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.XPATH, arg1))) #10초동안 해당 엘리먼트가 존재하는지 대기
###############################
# << 메인 >>
###############################
if __name__ == "__main__":
    # 사용자 입력
    sabun = pyautogui.prompt("로그인 사번을 입력해주세요.","입력") #사용자 입력 사번
    password = pyautogui.password("로그인 비밀번호를 입력해주세요.","입력") #사용자 입력 패스워드
    
    # browser 셋팅
    browser = webdriver.Chrome(executable_path='/Users/chaekyunghoon/desktop/PythonWorkSpace/chromedriver')
    url = "https://hmp.hanwhalife.com/online/fp" #FP월드 운영 URL
    browser.get(url) #FP월드 URL
    
    # 엑셀 활성화
    wb = load_workbook("FPworld_testCase.xlsx") #FPworld_testCase.xlsx 파일에서 wb를 불러옴
    ws = wb.active  # 현재 활성화된 sheet 가져옴

    login() # 로그인 화면 처리
    popUpClose() #팝업 종료 함수

    ###############################
    # 테스트 건수 만큼 반복
    ###############################
    for i in range(3,ws.max_row+1):
        goodFind()                      # 엑셀 상품명과 일치하는 메뉴명 찾기
        pyautogui.press('enter')        # alert창 처리
        time.sleep(1)
        mainSelect()                    # 상품 화면으로 프레임 전환
        setInput('ibxFpNo', 3)          # FP사번 set
        setInput('ibxCustName', 4)      # 고객명 set
        setOccpCode()                   # 직종정보 set(직종정보를 변경하면 보험종류 및나머지 계약자 항목이 리셋되어 제일 먼저 처리)
        setInsrKind()                   # 보험종류 set
        setCtorTable()                  # 계약자정보 set
        setBasicMainPlan()              # 기본설계 주계약 set    
        setRetirePlan()                 # 연금설계 set
        setAddPay()                     # 추가설계_추가납입 set
        setDraw()                       # 추가설계_인출 set
        setSmartChng()                  # 스마트전환형 set
        setSmartAddPay()                # 스마트전환형_추가납입 set
        setPreRetireAmt()               # 연금선지급 set
        setGuRtChngObjt()               # 보장전환형_대상 set
        setGuRtRetireChng()             # 보장전환형_연금설계 set
        setGuRtAddPay()                 # 보장전환형_추가납입 set
        setBasicAddPlan()               # 기본설계_부가특약 set
        joinSave()                      # 가입설계 저장
        
    wb.close()
    pyautogui.alert('가입설계 테스트 자동화 종료')