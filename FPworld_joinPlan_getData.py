import pyautogui
import pyperclip
import time
import datetime
import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from datetime import datetime, timedelta      
from bs4 import BeautifulSoup  
from selenium.common.exceptions import NoSuchElementException

###############################
# 엑셀 항목 순서 변수
###############################
insrKindNo      = 6       # 보험종류
ctorInfoNo      = 10      # 계약자 정보
basicMainPlanNo = 35      # 기본설계 주계약
###############################
# 로그인 화면 처리
###############################
def login():
    browser.maximize_window()
    time.sleep(1)
    browser.find_element(By.ID, 'ibxUserId').send_keys(sabun)
    time.sleep(0.5)
    browser.find_element(By.ID, 'ibxPassword').send_keys(password)
    time.sleep(0.5)
    browser.find_element(By.ID, 'rdoTitl_input_0').click() #스마트인증 라디오 버튼
    time.sleep(0.5)
    browser.find_element(By.XPATH, '//*[@id="bt_login"]/a').click() #로그인 버튼
    time.sleep(0.5)
###############################
# 팝업 종료 처리
###############################
def popUpClose():
    browser.implicitly_wait(10)
    time.sleep(1)
    handles = browser.window_handles #메인페이지 및 팝업 갯수

    # 메인페이지가 아니면 팝업(브라우저) 종료
    for handle in handles:
        if handle != handles[0]:
            browser.switch_to.window(handle)
            browser.close()

    time.sleep(0.5)
    browser.switch_to.window(browser.window_handles[0]) #메인화면으로 전환
    elem = browser.find_element(By.XPATH, '//*[@id="smallpop_close"]') #내부팝업(?) 종료
    browser.execute_script("arguments[0].click();", elem)
    time.sleep(0.5)
###############################
# 엑셀 상품명과 일치하는 메뉴명 찾기
###############################
def goodFind():
    browser.switch_to.default_content() #메인화면으로 전환
    browser.find_element(By.ID, 'genMenuDepth1_3_menuNm').click()
    elems = browser.find_elements(By.CLASS_NAME, 'w2anchor2 ') #메뉴명 전체 가져오기
    for elem in elems:
        if elem.text == goodName: 
            time.sleep(0.5)
            elem.click()
            print('### ' + str(i) + '. 상품명 : ' + str(goodName))
            break
###############################
# 페이지 로딩 대기
###############################
def pageWait(arg1):
    elem = WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.XPATH, arg1))) #10초동안 해당 엘리먼트가 존재하는지 대기
###############################
# 설계이력 조회
###############################
def history():
    #frame 전환
    browser.implicitly_wait(10)
    browser.switch_to.frame('windowContainer1_subWindow0_iframe') 
    time.sleep(0.5)

    # FP사번 입력
    fpNo = ws.cell(row=i,column=3).value
    elem = browser.find_element(By.ID, 'ibxFpNo')
    time.sleep(1)
    elem.clear()
    time.sleep(1)
    elem.send_keys(fpNo)
    time.sleep(0.5)

    # 엔터키(조회)
    pyautogui.press('enter') 
    time.sleep(1)

    #메인 조회 페이지 로딩 대기
    pageWait('//*[@id="btnPlanSearch"]/a')

    #설계이력조회
    browser.find_element(By.XPATH, '//*[@id="btnPlanSearch"]/a').click()
    browser.switch_to.frame('ncjsi408pvw_iframe') #frame 전환
    time.sleep(0.5)

    # 고객명 입력
    custName = ws.cell(row=i,column=4).value
    browser.find_element(By.ID, 'ibxCust').send_keys(custName)
    
    # 시작일자 input값 가져오기
    elem = browser.find_element(By.ID, 'icaFrDate_input') 
    strDate = elem.get_attribute("value")
    
    # 문자열 날자형식으로 변환 후 7일전 값으로 입력
    date = datetime.strptime(strDate.replace('-',''),'%Y%m%d').date()
    date = date + timedelta(days=-7)
    elem.clear()
    elem.send_keys(str(date))
    time.sleep(0.5)
    
    # 조회버튼
    browser.find_element(By.XPATH, '//*[@id="btnSearch"]/a').click() 
    time.sleep(3)
    
    browser.implicitly_wait(10)
    pageWait('//*[@id="grdIndList_cell_0_3"]/nobr') # 설계이력조회 페이지 로딩 대기

    # 엑셀에 입력된 설계번호 클릭
    elems = browser.find_elements(By.TAG_NAME, 'a')
    for elem in elems:
        if str(elem.text) == str(ws.cell(row=i,column=5).value):
            elem.click()
###############################
# 메인 조회
###############################          
def mainSelect():
    # 메인frame으로 전환
    time.sleep(1)
    browser.switch_to.default_content() #메인화면으로 전환
    time.sleep(0.5)
    browser.switch_to.frame('windowContainer1_subWindow0_iframe') #frame 전환

    #페이지 로딩 대기
    time.sleep(3)
    browser.implicitly_wait(10)
    pageWait('//*[@id="sbxMenu0_label"]') #메인 조회 페이지 로딩 대기
###############################
# 보험종류 가져오기
###############################
def getInsrKind():
    browser.implicitly_wait(0.5)
    try:
        print("### 보험종류 가져오기 getInsrKind 시작 ###")
        for k in range(0,4):
            elem = browser.find_element(By.ID, 'sbxMenu'+str(k)+'_label')
            if str(elem.text):
                ws.cell(row=i,column=insrKindNo+k).value = str(elem.text)
    except Exception as e:
        print("### 보험종류 가져오기 getInsrKind 예외발생: " + str(e))
        pass
    
    print("### 보험종류 가져오기 getInsrKind 종료 ###")
    time.sleep(0.5)
###############################
# 계약자 구분 테이블 가져오기
###############################
def getCtorTable():
    print("### 계약자 구분 getCtorTable 시작 ###")
    browser.implicitly_wait(0.5)
    try:
        elems = browser.find_elements(By.TAG_NAME, "nobr")
        for elem in elems:
            if str(elem.text):
                if '보험계약자' == str(elem.text):
                    getCtorInfo(0,ctorInfoNo) # 계약자 정보 가져오기 함수 호출 arg1: xpath, arg2: i
                elif '피보험자(주피)' == str(elem.text):
                    getCtorInfo(1,ctorInfoNo+5) # 계약자 정보 가져오기 함수 호출
                elif '배우자(종피)' == str(elem.text):
                    getCtorInfo(2,ctorInfoNo+10) # 계약자 정보 가져오기 함수 호출
                elif '자녀1' == str(elem.text):
                    getCtorInfo(3,ctorInfoNo+15) # 계약자 정보 가져오기 함수 호출
                elif '자녀2' == str(elem.text):
                    getCtorInfo(4,ctorInfoNo+20) # 계약자 정보 가져오기 함수 호출
    except Exception as e:
        print("### 계약자 구분 getCtorTable 예외발생: " + str(e))
        pass
    print("### 계약자 구분 getCtorTable 종료 ###")
    time.sleep(0.5)
###############################
# 계약자정보 가져오기
###############################
def getCtorInfo(arg1,arg2):
    browser.implicitly_wait(0.5)
    try:
        print("### 계약자정보 getCtorInfo 시작 ###")
        # 이름
        ctorName = browser.find_element(By.XPATH, '//*[@id="grdInsu_cell_'+str(arg1)+'_1"]/nobr')
        ws.cell(row=i,column=arg2).value = ctorName.text
        # 건강체
        healYn = browser.find_element(By.XPATH, '//*[@id="grdInsu_cell_'+str(arg1)+'_8"]/input')
        if healYn.is_selected() == "True":
            ws.cell(row=i,column=arg2+1).value = "Y"
        else:
            ws.cell(row=i,column=arg2+1).value = "N"
        # 직종코드
        occpCode = browser.find_element(By.XPATH, '//*[@id="grdInsu_cell_'+str(arg1)+'_11"]/nobr')
        ws.cell(row=i,column=arg2+2).value = occpCode.text
        # 외국인
        frnr = browser.find_element(By.XPATH, '//*[@id="grdInsu_cell_'+str(arg1)+'_15"]/nobr')
        ws.cell(row=i,column=arg2+3).value = frnr.text
        # 체류
        stay = browser.find_element(By.XPATH, '//*[@id="grdInsu_cell_'+str(arg1)+'_16"]/nobr')
        ws.cell(row=i,column=arg2+4).value = stay.text 
        print("### 계약자정보 getCtorInfo 종료 ###")
    except Exception as e:
        print("### 계약자정보 getCtorInfo 예외발생: " + str(e))
        pass
    time.sleep(0.5)
###############################
# 기본설계 주계약 가져오기
###############################
def getBasicMainPlan():
    browser.implicitly_wait(0.5)
    try:
        print("### 기본설계 주계약 getBasicMainPlan 시작 ###")
        global currentCol
        elem = browser.find_element(By.XPATH ,'//*[@id="grdMain_cell_0_1"]/nobr') #합계보험료
        ws.cell(row=i,column=basicMainPlanNo).value = elem.text.replace(",",'')
        elem = browser.find_element(By.XPATH ,'//*[@id="grdMain_cell_0_4"]/nobr') #가입금액
        ws.cell(row=i,column=basicMainPlanNo+1).value = elem.text.replace(",",'')
        elem = browser.find_element(By.XPATH ,'//*[@id="grdMain_cell_0_7"]/nobr') #보험기간
        ws.cell(row=i,column=basicMainPlanNo+2).value = elem.text.replace(",",'')
        elem = browser.find_element(By.XPATH ,'//*[@id="grdMain_cell_0_9"]/nobr') #납입기간
        ws.cell(row=i,column=basicMainPlanNo+3).value = elem.text.replace(",",'')
        elem = browser.find_element(By.XPATH ,'//*[@id="grdMain_cell_0_11"]/nobr') #납입주기
        ws.cell(row=i,column=basicMainPlanNo+4).value = elem.text.replace(",",'')
        elem = browser.find_element(By.XPATH ,'//*[@id="grdMain_cell_0_13"]/nobr') #증액보험기간
        ws.cell(row=i,column=basicMainPlanNo+5).value = elem.text.replace(",",'')
    except NoSuchElementException:
        print("### 기본설계 주계약 getBasicMainPlan 데이터 없음")
        pass
    currentCol = basicMainPlanNo+5 ##현재 엑셀 컬럼 위치
    print("### 기본설계 주계약 getBasicMainPlan 종료 ###")
    time.sleep(0.5)
###############################
# 연금설계 가져오기
###############################
def getRetirePlan():
    print("### 연금설계 getRetirePlan 시작 ###")
    browser.implicitly_wait(0.5)
    try:
        global currentCol
        # 엑셀 항목명 셋팅
        ws.cell(row=1,column=currentCol+1).value = '연금설계'
        ws.cell(row=2,column=currentCol+1).value = '개시나이'
        ws.cell(row=2,column=currentCol+2).value = '예시나이간격'
        # 연금설계탭
        elem = browser.find_element(By.XPATH, '//*[@id="subTac_tab_tab02"]/div[1]/a') 
        elem.click()
        # 데이터 스크래핑
        elem = browser.find_element(By.XPATH ,'//*[@id="sbxAnntStrtAge_label"]') #연금개시나이
        ws.cell(row=i,column=currentCol+1).value = elem.text.replace("-선택-",'')
        elem = browser.find_element(By.XPATH ,'//*[@id="sbxAnntAgeTerm_label"]') #예시나이간격
        ws.cell(row=i,column=currentCol+2).value = elem.text.replace("-선택-",'') 
    except Exception as e:
        print("### 연금설계 getRetirePlan 예외발생: " + str(e))
        pass

    # 현재 엑셀 컬럼 위치
    currentCol = currentCol+2
    wb.save("FPworld_testCase.xlsx")
    print("### 연금설계 getRetirePlan 종료 ###")
    time.sleep(0.5)
###############################
# 추가설계_추가납입 가져오기
###############################
def getAddPay():
    print("### 추가설계_추가납입 getAddPay 시작 ###")
    browser.implicitly_wait(0.5)
    try:
        global currentCol
        currentColTemp = currentCol

        try:
            # 팝업창 처리
            WebDriverWait(browser, 3).until(EC.alert_is_present())
            alert = browser.switch_to.alert
            alert.accept() 
        # 팝업창이 없으면
        except:
            pass
        
        k=0
        # 엑셀 항목명 셋팅
        for row in range(0,5):
            ws.cell(row=1,column=currentCol+1+k).value = '추가설계_추가납입'
            ws.cell(row=2,column=currentCol+1+k).value = '추가납입시점'
            ws.cell(row=2,column=currentCol+2+k).value = '추가납입기간'
            ws.cell(row=2,column=currentCol+3+k).value = '추가납입금액_주기'
            ws.cell(row=2,column=currentCol+4+k).value = '추가납입금액'
            currentCol = currentCol+4+k
            currentColFinal = currentCol # 엑셀 데이터 마지막 위치

        # 추가설계탭
        elem = browser.find_element(By.XPATH, '//*[@id="subTac_tab_tab03"]/div[1]/a') 
        elem.click()
        currentCol = currentColTemp
        k=0
        for row in range(0,5):
            # 추가납입 col 데이터
            for col in ['0','2','4','5']:        
                # 데이터 스크래핑
                elem = browser.find_element(By.XPATH ,"//*[@id='grdGurtAddPay_cell_" + str(row) + "_" + str(col) + "']/nobr") 
                ws.cell(row=i,column=currentCol+1+k).value = elem.text.replace("선택",'')
                k = k +1
    except Exception as e:
        print("### 추가설계_추가납입 getAddPay 예외발생: " + str(e))
        pass

    currentCol = currentColFinal    
    wb.save("FPworld_testCase.xlsx")
    print("### 추가설계_추가납입 getAddPay 종료 ###")
    time.sleep(0.5)
###############################
# 추가설계_인출 가져오기
###############################
def getDraw():
    print("### 추가설계_인출 getDraw 시작 ###")
    browser.implicitly_wait(0.5)
    try:
        global currentCol
        currentColTemp = currentCol

        try:
            # 팝업창 처리
            WebDriverWait(browser, 3).until(EC.alert_is_present())
            alert = browser.switch_to.alert
            alert.accept() 
        # 팝업창이 없으면
        except:
            pass

        k=0
        # 인출
        for row in range(0,5):
            ws.cell(row=1,column=currentCol+1+k).value = '추가설계_인출'
            ws.cell(row=2,column=currentCol+1+k).value = '계약자적립금인출_용도'
            ws.cell(row=2,column=currentCol+2+k).value = '계약자적립금인출_시기(시작)'
            ws.cell(row=2,column=currentCol+3+k).value = '계약자적립금인출_시기(종료)'
            ws.cell(row=2,column=currentCol+4+k).value = '계약자적립금인출_금액'
            currentCol = currentCol+4+k
            currentColFinal = currentCol # 엑셀 데이터 마지막 위치
        
        # 추가설계탭
        elem = browser.find_element(By.XPATH, '//*[@id="subTac_tab_tab03"]/div[1]/a') 
        elem.click()
        currentCol = currentColTemp
        k=0
        # 인출
        for row in range(0,5):
            # 인출 col 데이터
            for col in ['0','1','4','6']:
                # 데이터 스크래핑
                elem = browser.find_element(By.XPATH ,"//*[@id='grdGurtDraw_cell_" + str(row) + "_" + str(col) + "']/nobr")
                ws.cell(row=i,column=currentCol+1+k).value = elem.text.replace("선택",'')
                k = k +1       
    except Exception as e:
        print("### 추가설계_인출 getDraw 예외발생: " + str(e))
        pass

    currentCol = currentColFinal 
    wb.save("FPworld_testCase.xlsx")
    print("### 추가설계_인출 getDraw 종료 ###")
    time.sleep(0.5)
###############################
# 스마트전환형 가져오기
###############################
def getSmartChng():
    print("### 스마트전환형 getSmartChng 시작 ###")
    browser.implicitly_wait(0.5)
    try:
        global currentCol
        try:
            # 팝업창 처리
            WebDriverWait(browser, 3).until(EC.alert_is_present())
            alert = browser.switch_to.alert
            alert.accept() 
        # 팝업창이 없으면
        except:
            pass

        # 엑셀 항목명 셋팅
        ws.cell(row=1,column=currentCol+1).value = '스마트전환형'
        ws.cell(row=2,column=currentCol+1).value = '스마트전환대상자'
        ws.cell(row=2,column=currentCol+2).value = '스마트전환시점'
        ws.cell(row=2,column=currentCol+3).value = '보험기간'
        ws.cell(row=2,column=currentCol+4).value = '납입기간'
        # 스마트전환형탭
        elem = browser.find_element(By.XPATH, '//*[@id="subTac_tab_tab05"]/div[1]/a') #스마트전환형탭
        if '스마트전환형' == elem.text: 
            elem.click()
        # 데이터 스크래핑
        elem = browser.find_element(By.XPATH ,'//*[@id="sbxConv8GurtConv_label"]') #스마트전환 대상자
        ws.cell(row=i,column=currentCol+1).value = elem.text.replace("-선택-",'')
        elem = browser.find_element(By.XPATH ,'//*[@id="sbxConv1GurtConv_label"]') #스마트전환 시점
        ws.cell(row=i,column=currentCol+2).value = elem.text.replace("-선택-",'')
        elem = browser.find_element(By.XPATH ,'//*[@id="sbxConv4GurtConv_label"]') #보험기간
        ws.cell(row=i,column=currentCol+3).value = elem.text.replace("-선택-",'')
        elem = browser.find_element(By.XPATH ,'//*[@id="sbxConv5GurtConv_label"]') #납입기간
        ws.cell(row=i,column=currentCol+4).value = elem.text.replace("-선택-",'')
    except Exception as e:
        print("### 스마트전환형 getSmartChng 예외발생: " + str(e))
        pass

    currentCol = currentCol+4 #현재 엑셀 컬럼 위치
    wb.save("FPworld_testCase.xlsx")
    print("### 스마트전환형 getSmartChng 종료 ###")
    time.sleep(0.5)
###############################
# 스마트전환형_추가납입 가져오기
###############################
def getSmartAddPay():
    print("### 스마트전환형_추가납입 getSmartAddPay 시작 ###")
    browser.implicitly_wait(0.5)
    try:
        global currentCol
        currentColTemp = currentCol
        k=0
        # 엑셀 항목명 셋팅
        for row in range(0,5):
            ws.cell(row=1,column=currentCol+1+k).value = '스마트전환형_추가납입'
            ws.cell(row=2,column=currentCol+1+k).value = '추가납입시점'
            ws.cell(row=2,column=currentCol+2+k).value = '추가납입기간'
            ws.cell(row=2,column=currentCol+3+k).value = '추가납입금액_주기'
            ws.cell(row=2,column=currentCol+4+k).value = '추가납입금액'
            currentCol = currentCol+4+k
            currentColFinal = currentCol # 엑셀 데이터 마지막 위치

        # 스마트전환형탭
        elem = browser.find_element(By.XPATH, '//*[@id="subTac_tab_tab05"]/div[1]/a') #스마트전환형탭
        if '스마트전환형' == elem.text: 
            elem.click()
        currentCol = currentColTemp
        k=0
        for row in range(0,5):
            # 추가납입 col 데이터
            for col in ['0','2','4','5']:        
                # 데이터 스크래핑
                elem = browser.find_element(By.XPATH ,"//*[@id='grdGurtConvAddPay_cell_" + str(row) + "_" + str(col) + "']/nobr") 
                ws.cell(row=i,column=currentCol+1+k).value = elem.text.replace("선택",'')
                k = k +1
    except Exception as e:
        print("### 스마트전환형_추가납입 getSmartAddPay 예외발생: " + str(e))
        pass

    currentCol = currentColFinal    
    wb.save("FPworld_testCase.xlsx")
    print("### 스마트전환형_추가납입 getSmartAddPay종료 ###")
    time.sleep(0.5)
###############################
# 연금선지급 가져오기
###############################
def getPreRetireAmt():
    print("### 연금선지급 getPreRetireAmt 시작 ###")
    browser.implicitly_wait(0.5)
    try:
        global currentCol
        # 엑셀 항목명 셋팅
        ws.cell(row=1,column=currentCol+1).value = '연금선지급'
        ws.cell(row=2,column=currentCol+1).value = '선지급나이'
        ws.cell(row=2,column=currentCol+2).value = '선지급비율'
        ws.cell(row=2,column=currentCol+3).value = '선지급기간'
        # 연금선지급탭
        elem = browser.find_element(By.XPATH, '//*[@id="subTac_tab_tab07"]/div[1]/a') 
        elem.click()
        # 데이터 스크래핑
        elem = browser.find_element(By.XPATH ,'//*[@id="sbxAnntAdvpymt1_label"]') #선지급나이
        ws.cell(row=i,column=currentCol+1).value = elem.text.replace("-선택-",'')
        elem = browser.find_element(By.XPATH ,'//*[@id="sbxAnntAdvpymt2_label"]') #선지급비율
        ws.cell(row=i,column=currentCol+2).value = elem.text.replace("-선택-",'')
        elem = browser.find_element(By.XPATH ,'//*[@id="sbxAnntAdvpymt3_label"]') #선지급기간
        ws.cell(row=i,column=currentCol+3).value = elem.text.replace("-선택-",'')
    except Exception as e:
        print("### 연금선지급 getPreRetireAmt 예외발생: " + str(e))
        pass

    currentCol = currentCol + 3 #현재 엑셀 컬럼 위치
    wb.save("FPworld_testCase.xlsx")
    print("### 연금선지급 getPreRetireAmt 종료 ###")
    time.sleep(0.5)
###############################
# 보장전환형_대상 가져오기
###############################
def getGuRtChngObjt():
    print("### 보장전환형_대상 getGuRtChng 시작 ###")
    browser.implicitly_wait(0.5)
    try:
        global currentCol
        # 엑셀 항목명 셋팅
        ws.cell(row=1,column=currentCol+1).value = '보장전환형_대상'
        ws.cell(row=2,column=currentCol+1).value = '보장전환대상자'
        ws.cell(row=2,column=currentCol+2).value = '보장전환시점'
        ws.cell(row=2,column=currentCol+3).value = '보험기간'
        ws.cell(row=2,column=currentCol+4).value = '납입기간'
        # 보장전환형탭
        elem = browser.find_element(By.XPATH, '//*[@id="subTac_tab_tab05"]/div[1]/a')
        if '보장전환형' == elem.text: 
            elem.click()
        # 데이터 스크래핑
        elem = browser.find_element(By.XPATH ,'//*[@id="sbxConv8GurtConv_label"]') #보험전환대상자
        ws.cell(row=i,column=currentCol+1).value = elem.text.replace("-선택-",'')
        elem = browser.find_element(By.XPATH ,'//*[@id="sbxConv1GurtConv_label"]') #보장전환시점
        ws.cell(row=i,column=currentCol+2).value = elem.text.replace("-선택-",'')
        elem = browser.find_element(By.XPATH ,'//*[@id="sbxConv4GurtConv_label"]') #보험기간
        ws.cell(row=i,column=currentCol+3).value = elem.text.replace("-선택-",'')
        elem = browser.find_element(By.XPATH ,'//*[@id="sbxConv5GurtConv_label"]') #납입기간
        ws.cell(row=i,column=currentCol+4).value = elem.text.replace("-선택-",'')
    except Exception as e:
        print("### 보장전환형_대상 getGuRtChng 예외발생: " + str(e))
        pass

    currentCol = currentCol + 4 #현재 엑셀 컬럼 위치
    wb.save("FPworld_testCase.xlsx")
    print("### 보장전환형_대상 getGuRtChng 종료 ###")
    time.sleep(0.5)
###############################
# 보장전환형_연금전환설계 가져오기
###############################
def getGuRtRetireChng():
    print("### 보장전환형_연금전환설계 getGuRtRetireChng 시작 ###")
    browser.implicitly_wait(0.5)
    try:
        global currentCol
        # 엑셀 항목명 셋팅
        ws.cell(row=1,column=currentCol+1).value = '보장전환형_연금전환설계'
        ws.cell(row=2,column=currentCol+1).value = '연금개시나이'
        ws.cell(row=2,column=currentCol+2).value = '예시나이간격'
        # 보장전환형탭
        elem = browser.find_element(By.XPATH, '//*[@id="subTac_tab_tab05"]/div[1]/a')
        if '보장전환형' == elem.text: 
            elem.click()
        # 데이터 스크래핑
        elem = browser.find_element(By.XPATH ,'//*[@id="sbxGurtConvAnntStrtAge_label"]') #연금개시나이
        ws.cell(row=i,column=currentCol+1).value = elem.text.replace("-선택-",'')
        elem = browser.find_element(By.XPATH ,'//*[@id="sbxGurtConvAnntAgeTerm_label"]') #예시나이간격
        ws.cell(row=i,column=currentCol+2).value = elem.text.replace("-선택-",'') 
    except Exception as e:
        print("### 보장전환형_연금전환설계 getGuRtRetireChng 예외발생: " + str(e))
        pass

    # 현재 엑셀 컬럼 위치
    currentCol = currentCol+2
    wb.save("FPworld_testCase.xlsx")
    print("### 보장전환형_연금전환설계 getGuRtRetireChng 종료 ###")
    time.sleep(0.5)
###############################
# 보장전환형_추가납입 가져오기
###############################
def getGuRtAddPay():
    print("### 보장전환형_추가납입 getGuRtAddPay 시작 ###")
    browser.implicitly_wait(0.5)
    try:
        global currentCol
        currentColTemp = currentCol
        k=0
        # 엑셀 항목명 셋팅
        for row in range(0,5):
            ws.cell(row=1,column=currentCol+1+k).value = '보장전환형_추가납입'
            ws.cell(row=2,column=currentCol+1+k).value = '추가납입시점'
            ws.cell(row=2,column=currentCol+2+k).value = '추가납입기간'
            ws.cell(row=2,column=currentCol+3+k).value = '추가납입금액_주기'
            ws.cell(row=2,column=currentCol+4+k).value = '추가납입금액'
            currentCol = currentCol+4+k
            currentColFinal = currentCol # 엑셀 데이터 마지막 위치

        # 보장전환형탭
        elem = browser.find_element(By.XPATH, '//*[@id="subTac_tab_tab05"]/div[1]/a') 
        if '보장전환형' == elem.text: 
            elem.click()
        currentCol = currentColTemp
        k=0
        for row in range(0,5):
            # 추가납입 col 데이터
            for col in ['0','2','4','5']: 
                # 데이터 스크래핑
                elem = browser.find_element(By.XPATH ,"//*[@id='grdGurtConvAddPay_cell_" + str(row) + "_" + str(col) + "']/nobr") 
                ws.cell(row=i,column=currentCol+1+k).value = elem.text.replace("선택",'')
                k = k +1
    except Exception as e:
        print("### 보장전환형_추가납입 getGuRtAddPay 예외발생: " + str(e))
        pass

    currentCol = currentColFinal    
    wb.save("FPworld_testCase.xlsx")
    print("### 보장전환형_추가납입 getGuRtAddPay 종료 ###")
    time.sleep(0.5)
###############################
# 기본설계_부가특약 가져오기
###############################
def getBasicAddPlan():
    browser.implicitly_wait(0.5)
    try:
        print("### 부가특약 getBasicAddPlan 시작 ### ")
        global currentCol
        elem = browser.find_element(By.XPATH, '//*[@id="subTac_tab_tab01"]/div[1]/a') #기본설계탭
        elem.click()

        # 부가특약 1~9라인 처리
        k=0
        for row in range(0,9):
            for col in range(0,5):
                elem = browser.find_element(By.XPATH ,"//*[@id='grdGoodMenu_cell_" + str(row) + "_" + str(4) + "']/nobr") #가입금액
                # 부가특약 가입금액이 0이 아니면
                if str(elem.text) != str(0): 
                    # 항목명 넣기
                    if col == 0:
                        ws.cell(row=1,column=currentCol+1+k).value = '부가특약'
                        ws.cell(row=2,column=currentCol+1+k).value = '구분'
                    elif col == 1:
                        ws.cell(row=2,column=currentCol+1+k).value = '특약명칭'
                    elif col == 2:
                        ws.cell(row=2,column=currentCol+1+k).value = '보험기간'
                    elif col == 3:
                        ws.cell(row=2,column=currentCol+1+k).value = '납입기간'
                    elif col == 4:
                        ws.cell(row=2,column=currentCol+1+k).value = '가입금액' 
                    #부가특약 정보 가져오기
                    elem = browser.find_element(By.XPATH ,"//*[@id='grdGoodMenu_cell_" + str(row) + "_" + str(col) + "']/nobr") 
                    ws.cell(row=i,column=currentCol+1+k).value = str(elem.text)
                    k = k +1
                    wb.save("FPworld_testCase.xlsx")
        #10번째 라인부터 동적스크롤 처리
        while True:
            # 현재 9번째 라인 특약명칭
            current_elem = browser.find_element(By.XPATH , '//*[@id="grdGoodMenu_cell_8_1"]/nobr') 
            current_elem.click()
            current9line = str(current_elem.text)

            # 작은 스크롤 아래로 한칸 이동
            current_elem.click()
            time.sleep(0.05)
            pyautogui.press('down') 
            time.sleep(0.05)

            # 동적스크롤 작동 후 9번째 라인 특약명칭
            next_elem = browser.find_element(By.XPATH , '//*[@id="grdGoodMenu_cell_8_1"]/nobr') 
            next9line = str(next_elem.text)

            if current9line == next9line: #현재 특약명과 다음 특약명이 일치하면 종료
                break
            else: # 현재 특약명과 다음 특약명이 불일치하면 엑셀에 저장
                # 부가특약 가입금액이 0이 아니면
                elem = browser.find_element(By.XPATH ,'//*[@id="grdGoodMenu_cell_8_4"]/nobr')
                if str(elem.text) != str(0):
                    for col in range(0,5):
                        # 항목명 넣기
                        if col == 0:
                            ws.cell(row=1,column=currentCol+1+k).value = '부가특약'
                            ws.cell(row=2,column=currentCol+1+k).value = '구분'
                        elif col == 1:
                            ws.cell(row=2,column=currentCol+1+k).value = '특약명칭'
                        elif col == 2:
                            ws.cell(row=2,column=currentCol+1+k).value = '보험기간'
                        elif col == 3:
                            ws.cell(row=2,column=currentCol+1+k).value = '납입기간'
                        elif col == 4:
                            ws.cell(row=2,column=currentCol+1+k).value = '가입금액'    
                        #부가특약 정보 가져오기
                        elem = browser.find_element(By.XPATH ,"//*[@id='grdGoodMenu_cell_8_" + str(col) + "']/nobr")
                        ws.cell(row=i,column=currentCol+1+k).value = str(elem.text)
                        k = k +1
                        wb.save("FPworld_testCase.xlsx")
    except Exception as e:
        print("### 부가특약 getBasicAddPlan 예외발생: " + str(e))
        pass
    wb.save("FPworld_testCase.xlsx")
    print("### 부가특약 getBasicAddPlan 종료 ### ")
    time.sleep(0.5)
###############################
# << 메인 >>
###############################
if __name__ == "__main__":
    # 사용자 입력
    sabun = pyautogui.prompt("로그인 사번을 입력해주세요.","입력") #사용자 입력 사번
    password = pyautogui.password("로그인 비밀번호를 입력해주세요.","입력") #사용자 입력 패스워드
    
    # browser 셋팅
    browser = webdriver.Chrome(executable_path='/Users/chaekyunghoon/desktop/PythonWorkSpace/chromedriver')
    url = "https://hmp.hanwhalife.com/online/fp" #FP월드 운영 URL
    browser.get(url) #FP월드 URL
    
    # 엑셀 활성화
    wb = load_workbook("FPworld_testCase.xlsx") #FPworld_testCase.xlsx 파일에서 wb를 불러옴
    ws = wb.active  # 현재 활성화된 sheet 가져옴

    login() # 로그인 화면 처리
    popUpClose() #팝업 종료 함수

    ###############################
    # 테스트 건수 만큼 반복
    ###############################
    for i in range(3,ws.max_row+1):
        currentCol = 0            # 현재 엑셀 컬럼 위치
        goodName = ws.cell(row=i,column=2).value #엑셀 상품명
        ws.cell(row=i,column=1).value = 'case' + str(i-2)

        goodFind() # 엑셀 상품명과 일치하는 메뉴명 찾기
        pyautogui.press('enter') # 팝업창 처리
        time.sleep(0.1)
        
        ###############################
        # 상품 공통
        ###############################
        history()          # 설계 이력조회
        mainSelect()       # 메인 조회
        getInsrKind()      # 보험종류
        getCtorTable()     # 계약자구분
        getBasicMainPlan() # 기본설계_주계약

        ###############################
        # 상품 옵션
        ###############################
        getRetirePlan()    # 연금설계
        print("1. 연금설계 종료: " + str(currentCol))
        getAddPay()        # 추가설계_추가납입
        print("2. 추가설계_추가납입 종료: " + str(currentCol))
        getDraw()          # 추가설계_인출
        print("3. 추가설계_인출 종료: " + str(currentCol))
        getSmartChng()     # 스마트전환형
        print("4. 스마트전환형 종료: " + str(currentCol))
        getSmartAddPay()   # 스마트전환형_추가납입
        print("5. 스마트전환형_추가납입 종료: " + str(currentCol))
        getPreRetireAmt()  # 연금선지급
        print("6. 연금선지급 종료: " + str(currentCol))
        getGuRtChngObjt()      # 보장전환형_대상
        print("7. 보장전환형_대상 종료: " + str(currentCol))
        getGuRtRetireChng()      # 보장전환형_연금전환설계
        print("8. 보장전환형_연금전환설계 종료: " + str(currentCol))
        getGuRtAddPay()      # 보장전환형_추가납입
        print("9. 보장전환형_추가납입 종료: " + str(currentCol))
        getBasicAddPlan()  # 기본설계_부가특약(최종)
            
    wb.save("FPworld_testCase.xlsx")
    wb.close()
    browser.quit()