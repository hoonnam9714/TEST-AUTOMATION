from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from openpyxl import workbook
from commonConst import *
from hmcTestData import hmcTestData # 엑셀전환을 위한 데이터 클래스
from typing import List

# 공통 함수를 정의 2021.11.22 이영준 작성

#=========================================================
# 함수명  : openWebDriver
# 기능   : 크롬 드라이버를 통해 Url Open & 생성된 WebDriver 반환
# 입력   : driverPath(크롬드라이버 path), Url(오픈할 페이지 url), emNo(접속시 사용되는 사번)
# 출력   : open 된 드라이버 값
#=========================================================
def openWebDriver(driverPath, targetUrl, emNo) :
    options = webdriver.ChromeOptions()
    options.add_experimental_option("excludeSwitches", ["enable-logging"])
    # options.add_argument("disable-gpu")

    if "hmcQaUrl" == targetUrl:
        url = hmcQaUrlPrefix+emNo+hmcQaUrlSuffix
    
    browser = webdriver.Chrome(executable_path=driverPath, options=options) #웹드라이버 설정(크롬)
    browser.get(url) 
    browser.maximize_window()
    browser.implicitly_wait(10)

    return browser

#=========================================================
# 함수명  : openExcelFile
# 기능   : openyxl을 사용 엑셀파일을 오픈 후 데이터 클래스로 반환한다.
# 입력   : filePath(파일 경로), filename(오픈할 파일명), fileLayout(레이아웃명)
# 출력   : open 된 파일의 활성화 되어있는 워크시트 값
#=========================================================
def openExcelFile(filePath, filename, dataType, title) :
    wb = load_workbook(filePath+filename)
    ws = wb.active

    dataFile: hmcTestData = []
    if "hmcTestData" == dataType:
        for x in range(title + 1 , ws.max_row + 1):
            data = hmcTestData() 
            # for y in range(1, ws.max_column + 1):
            #     data(y, ws.cell(row=x, column=y).value)
            #     ws._cells_by_row
            # dataFile.append(data)

            data.gbn             = ws.cell(row=x, column=1).value
            data.emNo            = ws.cell(row=x, column=2).value
            data.custNm          = ws.cell(row=x, column=3).value
            data.prdCd           = ws.cell(row=x, column=4).value
            data.prdNm           = ws.cell(row=x, column=5).value
            data.contrRrno       = ws.cell(row=x, column=6).value
            data.contrGndr       = ws.cell(row=x, column=7).value
            data.contrInsdSameYn = ws.cell(row=x, column=8).value
            data.insdRrno        = ws.cell(row=x, column=9).value
            data.insdGndr        = ws.cell(row=x, column=10).value 
            data.insdJobCode     = ws.cell(row=x, column=11).value 
            data.insd1Rrno       = ws.cell(row=x, column=12).value 
            data.insd1Gndr       = ws.cell(row=x, column=13).value 
            data.insd1JobCode    = ws.cell(row=x, column=14).value 
            data.insd2Rrno       = ws.cell(row=x, column=15).value 
            data.insd2Gndr       = ws.cell(row=x, column=16).value 
            data.insd2JobCode    = ws.cell(row=x, column=17).value 
            data.insd3Rrno       = ws.cell(row=x, column=18).value 
            data.insd3Gndr       = ws.cell(row=x, column=19).value 
            data.insd3JobCode    = ws.cell(row=x, column=20).value 
            data.insType1        = ws.cell(row=x, column=21).value 
            data.insType2        = ws.cell(row=x, column=22).value 
            data.insType3        = ws.cell(row=x, column=23).value 
            data.insType4        = ws.cell(row=x, column=24).value 
            data.insType5        = ws.cell(row=x, column=25).value 
            data.hbdyObjYn       = ws.cell(row=x, column=26).value 
            data.slctdPlan       = ws.cell(row=x, column=27).value 
            data.mnins           = ws.cell(row=x, column=28).value 
            data.insPrm          = ws.cell(row=x, column=29).value 
            data.antyOpenAge     = ws.cell(row=x, column=30).value 
            data.antyExAgeGap    = ws.cell(row=x, column=31).value 
            data.swtObjPerson    = ws.cell(row=x, column=32).value 
            data.swtTime         = ws.cell(row=x, column=33).value 
            data.swtInsPrd       = ws.cell(row=x, column=34).value 
            data.swtInsPmPrd     = ws.cell(row=x, column=35).value 
            data.antyPrePayAge   = ws.cell(row=x, column=36).value 
            data.antyPrePayRto   = ws.cell(row=x, column=37).value 
            data.antyPrePayPrd   = ws.cell(row=x, column=38).value 
            data.antyPrmPayAll   = ws.cell(row=x, column=39).value 
            data.antyPrmPayPrt   = ws.cell(row=x, column=40).value 
            data.antyPayTyp      = ws.cell(row=x, column=41).value 
            data.antyOpn         = ws.cell(row=x, column=42).value 
            data.antyFocusPrd    = ws.cell(row=x, column=43).value 
            data.insPrd          = ws.cell(row=x, column=44).value 
            data.pmPrd           = ws.cell(row=x, column=45).value 
            data.pmCyl           = ws.cell(row=x, column=46).value 
            data.rfndExcpYn      = ws.cell(row=x, column=47).value 
            data.pmExptExcpYn    = ws.cell(row=x, column=48).value 
            data.spcdCod1        = ws.cell(row=x, column=49).value 
            data.spcdNm1         = ws.cell(row=x, column=50).value 
            data.spcdAmt1        = ws.cell(row=x, column=51).value 
            data.spcdPrd1        = ws.cell(row=x, column=52).value 
            data.spcdPmPrd1      = ws.cell(row=x, column=53).value 
            data.spcdCod2        = ws.cell(row=x, column=54).value 
            data.spcdNm2         = ws.cell(row=x, column=55).value 
            data.spcdAmt2        = ws.cell(row=x, column=56).value 
            data.spcdPrd2        = ws.cell(row=x, column=57).value 
            data.spcdPmPrd2      = ws.cell(row=x, column=58).value 
            data.spcdCod3        = ws.cell(row=x, column=59).value 
            data.spcdNm3         = ws.cell(row=x, column=60).value 
            data.spcdAmt3        = ws.cell(row=x, column=61).value 
            data.spcdPrd3        = ws.cell(row=x, column=62).value 
            data.spcdPmPrd3      = ws.cell(row=x, column=63).value 
            data.spcdCod4        = ws.cell(row=x, column=64).value 
            data.spcdNm4         = ws.cell(row=x, column=65).value 
            data.spcdAmt4        = ws.cell(row=x, column=66).value 
            data.spcdPrd4        = ws.cell(row=x, column=67).value 
            data.spcdPmPrd4      = ws.cell(row=x, column=68).value 
            data.spcdCod5        = ws.cell(row=x, column=69).value 
            data.spcdNm5         = ws.cell(row=x, column=70).value 
            data.spcdAmt5        = ws.cell(row=x, column=71).value 
            data.spcdPrd5        = ws.cell(row=x, column=72).value 
            data.spcdPmPrd5      = ws.cell(row=x, column=73).value 
            data.etc             = ws.cell(row=x, column=74).value 
            dataFile.append(data)
        return dataFile

    return dataFile

################################################################
# 프로그램 시작점을 구분하기 위해 if __name__ == "__main__": 사용
################################################################
if __name__ == "__main__":
    print("당 소스가 돌아갈때만")