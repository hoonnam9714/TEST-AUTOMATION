# Generated by CHAE KYUNG HOON
import time
import datetime
import sys
import pyautogui
import pyperclip
import re
from PIL.ImageOps import grayscale
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait          #웹페이지 로딩 대기
from selenium.webdriver.support import expected_conditions as EC 
from openpyxl import load_workbook         #엑셀파일 불러오기
from openpyxl.drawing.image import Image   #엑셀 이미지 삽입
from selenium.webdriver.support.select import Select

###########################################
# 변수 (엘리먼트) - element class 로 구현하여 사용하는 것이 소스 코드 상으로 관리가 편합니다.
# '''
# class hmcWebElement:
#     hmc_xpath1 = 
# '''
###########################################
hmc_xpath1 = '//*[@id="app"]/div/section/div[2]/ul[2]/li/button'                              # 신규설계 버튼
hmc_xpath1_1 = '//*[@id="app"]/div/section/div[2]/ul[1]'                                      # 가입가능 상품 리스트
hmc_xpath2_0 =   '//*[@id="app"]/div/section/div[2]/div[2]/div/ul'                            # 상품군 선택
hmc_xpath2_4 = '//*[@id="container"]/div[1]/section/div[2]/div[1]/div/p/button'               # 보험종류변경
hmc_xpath2_5_0 = '//*[@id="container"]/div[2]/div[2]/div/ul/li'                               # 셀렉트박스
hmc_xpath2_6 = '//*[@id="container"]/div[2]/div[2]/footer/button'                             # 적용
hmc_xpath2_7 = '//*[@id="info-21"]/label'                                                     # 건강체 체크 박스
hmc_xpath2_8 = '//*[@id="container"]/div[3]/div[2]/div/ul/li'                                 # 플랜선택
hmc_xpath2_9 = '//*[@id="container"]/div[3]/div[2]/div/div/button/span'                       # 직접설계
hmc_xpath3 = '//*[@id="container"]/div[1]/footer/div/button'                                  # 다음 버튼
hmc_xpath4 = '//*[@id="container"]/div[1]/section/div[2]/div[2]/div[2]/h2'                    # 주계약
hmc_xpath5 = '//*[@id="container"]/div[1]/section/div[2]/div[2]/div[2]/div[1]/div/input'      # 주계약 가입금액
hmc_xpath5_1 = '//*[@id="container"]/div[5]/div[2]/div/div/button'                            # 특약전체선택
hmc_xpath5_1_1 = '//*[@id="container"]/div[4]/div[2]/div/div/button'                          # 연금특약전체선택
hmc_xpath5_2 = '//*[@id="container"]/div[5]/div[2]/div/footer/button'                         # 특약전체추가
hmc_xpath5_2_1 = '//*[@id="container"]/div[4]/div[2]/div/footer/button'                       # 연금특약전체추가
hmc_xpath5_3 = '//*[@id="container"]/div[1]/section/div[2]/div[2]/div[2]/div[2]/div[2]/select'# 납입기간
hmc_xpath5_4 = '//*[@id="container"]/div[1]/section/div[2]/div[2]/div[2]/ul/li[1]/label'      # 가입금액
hmc_xpath5_5 = '//*[@id="container"]/div[1]/section/div[2]/div[2]/div[2]/ul/li[2]/label'      # 합계보험료
hmc_xpath5_5_1 = '//*[@id="container"]/div[1]/section/div[2]/div[2]/div[2]/div[1]/div/input'  # #연금 통합형 보험료(거치)
hmc_xpath5_5_2 = '//*[@id="container"]/div[1]/section/div[2]/div[2]/div[2]/div[2]/div/input'  # #연금 통합형 보험료(적립)
hmc_xpath5_6 = '//*[@id="container"]/div[1]/section/div[2]/div[2]/div[2]/div[2]/div[1]/select'# 보험기간
hmc_xpath5_7   = '//*[@id="container"]/div[1]/section/div[2]/div[2]/div[2]/div[2]/div[3]/select'# 납입주기
hmc_xpath5_7_1 = '//*[@id="container"]/div[1]/section/div[2]/div[2]/div[2]/div[2]/div[5]/select'# 납입주기
hmc_xpath5_7_2 = '//*[@id="container"]/div[1]/section/div[2]/div[2]/div[2]/div[2]/div[6]/select'# 납입주기
hmc_xpath5_7_3 = '//*[@id="container"]/div[1]/section/div[2]/div[2]/div[2]/div[2]/div[7]/select'# 납입주기
hmc_xpath5_8 = '//*[@id="container"]/div[1]/section/div[2]/div[3]/ul/li/label'                # 환급특약제외, 연금납입면제특약제외
hmc_xpath6 = '//*[@id="container"]/div[1]/footer/div/button[2]'                               # 상세결과보기
hmc_xpath7 = '//*[@id="container"]/div[1]/section/div[2]/div[2]/div/div[2]/div/dl[1]/dd[1]'   # 계약자명
hmc_xpath8 = '/html/body/div[2]/div[1]/footer/div/button'                                     # 설계저장
hmc_xpath9 = '/html/body/div[2]/div/section/section/footer/div[1]/button'                     # 상품설명서 미리보기

###########################################
# select 옵션 선택
###########################################
def selectClick(arg1, arg2, arg3):
    for i in range(1,10):
        elem = browser.find_element_by_xpath(arg1+'['+str(i)+']')
        if str(ws.cell(row=2,column=arg2).value).strip() in elem.text.strip():
            elem.click()
            break

###########################################
# checkbox 선택
###########################################
def checkBoxClick(arg1, arg2):
    elem = browser.find_element_by_xpath(arg1)
    if "Y" == ws.cell(row=2,column=arg2).value:
        if elem.is_selected():
            print("이미 체크되어있음")
        else:
            browser.execute_script("arguments[0].click();", elem)                 #강제클릭으로 엘리먼트 에러 해결
    elif "N" == ws.cell(row=2,column=arg2).value:
        if elem.is_selected():
            browser.execute_script("arguments[0].click();", elem)                 #강제클릭으로 엘리먼트 에러 해결
        else:
            print("이미 체크되어있지 않음")

###########################################
# 메인 프로그램 시작
###########################################
wb = load_workbook("getInfo.xlsx") #testCase1.xlsx 파일에서 wb를 불러옴
ws = wb.active  # 현재 활성화된 sheet 가져옴

options = webdriver.ChromeOptions()
options.add_experimental_option("excludeSwitches", ["enable-logging"])
options.add_argument("disable-gpu")
browser = webdriver.Chrome(options=options)
    
# browser = webdriver.Chrome() #웹드라이버 설정(크롬)
hmc_url = "https://devhmc.hanwhalife.com:1080/sfa/incomeHmc?etrDvCd=01&token=aa2cb27ab1d84482aced7b5ee21c798f&sfaCmpnDvsn=01&mobUserPrno="+str(ws.cell(row=2,column=2).value)+"&mobUserDvsn=02&offcCode=00000" #HMC QA 테스트 URL
browser.get(hmc_url) #HMC QA 테스트 URL
browser.maximize_window()
browser.implicitly_wait(10)

#테스트 케이스 col건수 만큼 반복 수행
for y in range(2,ws.max_column + 1):
    if "고객명" in ws.cell(row=1,column=y).value: # 엑셀항목과 값 비교
        browser.implicitly_wait(10)
        time.sleep(0.5)
        browser.find_element(By.CSS_SELECTOR, ".search > input").send_keys(ws.cell(row=2,column=y).value) #엑셀에서 가져온 고객명 맵핑
        browser.find_element(By.CSS_SELECTOR, ".search > input").send_keys(Keys.ENTER)                    #고객명 엔터
        time.sleep(0.5)
        browser.find_element_by_xpath(hmc_xpath1).click()      #신규설계 버튼
    elif "상품명" in ws.cell(row=1,column=y).value:
        time.sleep(2)
        browser.implicitly_wait(10)
        goodName = ws.cell(row=2,column=y).value #엑셀상품코드
        j=0

        #상품 선택
        for i in range(1,5):            
            browser.implicitly_wait(10)
            elem = WebDriverWait(browser, 30).until(EC.presence_of_element_located((By.XPATH, hmc_xpath2_0+'/li['+str(i)+']/a'))) #엘리먼트 로딩 대기
            browser.execute_script("arguments[0].click();", elem)                 #강제클릭으로 엘리먼트 에러 해결
            
            browser.implicitly_wait(10)
            time.sleep(0.5)
            elems = browser.find_elements_by_class_name("p_name") #class_name "p_name"(상품명)을 가지는 모든 엘리먼트 가져오기
            
            existYn = "N"
            checkYn = "N"

            if goodName:
                for elem in elems:    #반복하면서 elems 하나씩 뽑아오기
                    if elem.text:
                        if elem.text == goodName:
                            elem.click() #엑셀파일에 상품명이 존재하고 화면값과 일치하면 클릭
                            existYn = "Y" 
                            break
            else:
                for elem in elems:    #반복하면서 elems 하나씩 뽑아오기
                    if elem.text:
                        ws.cell(row=2+j,column=4).value = elem.text #엑셀파일에 상품명이 존재하지 않으면 상품명 스크래핑
                        j = j+1
                        checkYn = "Y"

            if existYn == "Y": #엑셀파일에 상품명이 존재하고 화면값과 일치하면 클릭하고 반복문 종료(보험종류로 이동)
                break
        if checkYn == "Y":     #엑셀파일에 상품명이 존재하지 않으면 상품명 스크래핑 후 반복문 종료(프로그램 종료)
                break
    elif "보험종류1" == ws.cell(row=1,column=y).value:
        browser.implicitly_wait(10)
        time.sleep(0.5)
        pyautogui.moveTo(958,688) #지정한 위치로 마우스를 이동(팝업확인클릭)
        pyautogui.click()
        time.sleep(0.5)
    
        try:
            browser.find_element_by_xpath(hmc_xpath2_4).click()      #보험종류변경
        except:
            print("보험종류 변경없음")
        time.sleep(0.5)         

        
        elems1 = browser.find_elements_by_class_name("insureSelect") #class_name "insureSelect"(보험종류)을 가지는 모든 엘리먼트 가져오기
        count=0
        countDetail = []
        #보험종류 및 보험종류별 상세 갯수
        for e1 in elems1:    
            if e1.text:
                w = str(e1.text)
                countDetail.append(str(w.count('\n')))  #보험종류별 상세 갯수
                count = count+1   #보험종류 갯수
        checkYn = "N"
        for i in range(0,count): #보험종류 만큼 반복(col)
            for j in range(0,int(countDetail[i])): #보험종류별 상세 만큼 반복(row)
                elem = browser.find_element_by_xpath(hmc_xpath2_5_0+'['+str(i+1)+']/p/select/option'+'['+str(j+1)+']')
                time.sleep(0.5)
                if ws.cell(row=2+j,column=y+i).value:  #보험종류1이 존재하면 맵핑 후 다음으로 진행
                    elemText = str(ws.cell(row=2,column=y+i).value)
                    elem1 = browser.find_element_by_xpath(hmc_xpath2_5_0+'['+str(i+1)+']/p/select')
                    elem1.send_keys(elemText) #엑셀 보험종류로 select 값 셋팅
                else: #보험종류1이 빈값이면 보험종류 스크래핑 후 엑셀에 저장
                    ws.cell(row=j+2,column=y+i).value = elem.text #엑셀에 보험종류 값 채우기
                    checkYn = "Y"

        if checkYn == "Y": #보험종류만 채우고 종료
            break

        time.sleep(0.5)
        browser.find_element_by_xpath(hmc_xpath2_6).click()      #적용 버튼    
        time.sleep(0.5)
        browser.find_element_by_xpath(hmc_xpath3).click()        #다음 버튼    
        
        #플랜설계 화면 닫기 처리(직접설계)
        browser.implicitly_wait(10)
        time.sleep(3)
        time.sleep(0.5)
        pyautogui.moveTo(940,276) # 플랜설계화면 직접설계 버튼1
        pyautogui.click()  
        time.sleep(0.5)
        pyautogui.moveTo(940,330) # 플랜설계화면 직접설계 버튼2
        pyautogui.click()  
        time.sleep(0.5)

        element = browser.find_element(By.CSS_SELECTOR, ".b_plus:nth-child(3)")  #특약추가버튼
        browser.execute_script("arguments[0].click();", element)                 #강제클릭으로 엘리먼트 에러 해결
        time.sleep(0.5)
        browser.implicitly_wait(10)
        browser.find_element_by_class_name("btn_white_normal").click()        #특약전체선택1
        time.sleep(0.5)
        browser.find_element_by_class_name("btn_point_pop").click()        #특약전체추가1
        time.sleep(0.5)
    elif "특약코드" in ws.cell(row=1,column=y).value:
        rowCount=0  #row 숫자를 세기 위한 변수
        optionCodes = [] #특약코드를 담을 리스트
        optionNames = [] #특약명을 담을 리스트
        elems = browser.find_elements_by_tag_name("li") #화면의 li태그를 모두 가져옴
        for elem in elems:    #반복하면서 elems 하나씩 뽑아오기
            if elem.get_attribute("id"): 
                 rowCount = rowCount + 1
                 
                 optionCode = elem.get_attribute("id") #화면의 li태그중 id속성을 가지고 있는걸 모두 가져옴
                 optionCodes.append(optionCode) #특약코드 리스트

                 optionElem = browser.find_element_by_xpath('//*[@id='+optionCode+']/div/p[1]')
                 optionName = optionElem.text
                 optionNames.append(optionName) #특약명 리스트

        for i in range(0,2): #엑셀에 특약코드/명 입력
            for j in range(0,len(optionCodes)):
                ws.cell(row=2+j,column=y).value = optionCodes[j]
                ws.cell(row=2+j,column=y+1).value = optionNames[j]

    #마지막 col일 경우 종료
    elif "비고" in ws.cell(row=1,column=y).value:
        break
###########################################
#브라우저 종료
###########################################
browser.quit()
wb.save("getInfo.xlsx")
wb.close()
